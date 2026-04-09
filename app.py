import os
import base64
import json
import uuid
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from io import BytesIO

from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from dotenv import load_dotenv
from pydantic import BaseModel
from PIL import Image
from PIL.PngImagePlugin import PngInfo
import piexif
from piexif import TYPES
import anthropic
import urllib.request
import urllib.error
from bs4 import BeautifulSoup
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

load_dotenv()

app = FastAPI(title="GroundSwell℠ Image Metadata Tool")

# Configure paths
BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
PROCESSED_DIR = BASE_DIR / "processed"
TEMPLATES_DIR = BASE_DIR / "templates"

UPLOAD_DIR.mkdir(exist_ok=True)
PROCESSED_DIR.mkdir(exist_ok=True)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
app.mount("/processed", StaticFiles(directory=str(PROCESSED_DIR)), name="processed")

# Initialize Anthropic client
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

# GroundSwell metadata template based on historical data
GROUNDSWELL_CONTEXT = """
You are a metadata specialist for GroundSwell℠, a Business Ownership Platform company.

GroundSwell℠ focuses on:
- Organizational Alpha (combined incremental returns and decreased risks)
- Fractal Flywheel of Organizational Development℠
- Fractal Flywheel of Management & Leadership Development℠
- Changing Pains℠ (growing pains when organization's systems don't support its size)
- Stages of Organizational Development℠
- Business Ownership School℠
- Inner Optimization℠
- Pyramid of Tech Stack Development℠
- Direct Ownership


Common keywords include: equity partners, strategic partners, return on investment, business building partner,
investment partner, Groundswell, organizational strategies, business growth, risk management, performance enhancement,
business efficiency, strategic management, value creation, operational efficiency, fractal flywheel, flywheel,
business improvement, business advice, business partner, business valuation, private equity

Categories: Business Ownership Platform, Direct Ownership, Organizational Development, Management & Leadership Development

Contact info:
- Website: www.groundswell.co
- Phone: 435-214-2997
- Credit: GroundSwell
- Copyright Status: Protected
"""

METADATA_PROMPT = """
Based on the image filename and visual content, generate metadata for this GroundSwell℠ image.

Filename: {filename}

Generate the following metadata in JSON format:
{{
    "title": "The title/name of the concept shown (from filename or image)",
    "headline": "A brief compelling headline (1 sentence) describing the value proposition",
    "description": "A detailed 2-3 sentence description of the concept",
    "keywords": ["list", "of", "relevant", "keywords", "comma", "separated"],
    "category": "Business Ownership Platform",
    "supplemental_category": "The specific sub-category (e.g., Organizational Development, Management & Leadership Development, Direct Ownership, etc.)",
    "create_date": "{current_date}"
}}

Important:
- The title should match the concept name from the filename (remove file extension, clean up formatting)
- The headline should convey the business value
- The description should explain what the concept means and its importance to businesses
- Keywords should include: the concept name, GroundSwell, relevant business terms
- Base your response on the visual content and the GroundSwell context provided

{context}

Return ONLY the JSON object, no additional text.
"""

REFERENCE_CONTEXT_PROMPT = """

Additional Context Reference:
The following content was provided as additional context for this image. Use it to generate more accurate, detailed, and contextually relevant metadata:

---
{reference_content}
---

Incorporate relevant information from this reference into the title, headline, description, and keywords.
"""

MAX_REFERENCE_LENGTH = 5000


def fetch_url_content(url: str) -> str:
    """Fetch and extract text content from a URL."""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            html = response.read().decode('utf-8', errors='ignore')
        soup = BeautifulSoup(html, 'html.parser')
        for element in soup(['script', 'style', 'nav', 'footer']):
            element.decompose()
        text = soup.get_text(separator='\n', strip=True)
        return text[:MAX_REFERENCE_LENGTH]
    except Exception:
        return ""


def extract_file_text(file_data: bytes, filename: str) -> str:
    """Extract text from uploaded file based on extension."""
    ext = Path(filename).suffix.lower()
    try:
        if ext == '.pdf':
            reader = PdfReader(BytesIO(file_data))
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
                if len(text) > MAX_REFERENCE_LENGTH:
                    break
            return text[:MAX_REFERENCE_LENGTH]
        elif ext == '.docx':
            doc = DocxDocument(BytesIO(file_data))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text[:MAX_REFERENCE_LENGTH]
        elif ext in ('.xlsx', '.xls'):
            wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ", ".join(str(cell) for cell in row if cell is not None)
                    if row_text:
                        text += row_text + "\n"
                    if len(text) > MAX_REFERENCE_LENGTH:
                        break
                if len(text) > MAX_REFERENCE_LENGTH:
                    break
            wb.close()
            return text[:MAX_REFERENCE_LENGTH]
        elif ext in ('.txt', '.csv', '.md', '.json'):
            text = file_data.decode('utf-8', errors='ignore')
            return text[:MAX_REFERENCE_LENGTH]
        else:
            return ""
    except Exception:
        return ""


VARIANT_DIMENSIONS = {
    "content": None,
    "social": (1200, 720),
    "featured": (700, 400),
    "thumbnail": (232, 245),
}


def resize_image_to_fit(image_data: bytes, max_width: int, max_height: int, file_ext: str) -> bytes:
    """Resize image to fit within max_width x max_height, preserving aspect ratio."""
    img = Image.open(BytesIO(image_data))
    if img.width <= max_width and img.height <= max_height:
        return image_data
    img.thumbnail((max_width, max_height), Image.LANCZOS)
    output = BytesIO()
    if file_ext.lower() in ['.jpg', '.jpeg']:
        img.save(output, 'JPEG', quality=95)
    else:
        img.save(output, 'PNG')
    return output.getvalue()


def generate_download_filename(original_filename: str, variant_type: str, date_str: str) -> str:
    """Generate standardized filename: {date}-{cleaned-name}-{type}-w.{ext}"""
    stem = Path(original_filename).stem
    ext = Path(original_filename).suffix.lower()
    cleaned = re.sub(r'[^a-z0-9]+', '-', stem.lower()).strip('-')
    if not date_str:
        date_str = datetime.now().strftime('%Y.%m.%d')
    date_dotted = date_str.replace('-', '.')
    return f"{date_dotted}-{cleaned}-{variant_type}-w{ext}"


class SaveMetadataRequest(BaseModel):
    file_id: str
    metadata: dict


def encode_image_to_base64(file_path: Path) -> str:
    """Encode image file to base64 string."""
    with open(file_path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8")


def get_media_type(filename: str) -> str:
    """Get the media type based on file extension."""
    ext = Path(filename).suffix.lower()
    media_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp"
    }
    return media_types.get(ext, "image/jpeg")


async def analyze_image_with_claude(file_path: Path, filename: str, reference_context: str = "") -> dict:
    """Use Claude to analyze the image and generate metadata."""

    # Encode image
    image_data = encode_image_to_base64(file_path)
    media_type = get_media_type(filename)
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Create prompt
    prompt = METADATA_PROMPT.format(
        filename=filename,
        current_date=current_date,
        context=GROUNDSWELL_CONTEXT
    )

    if reference_context:
        prompt += REFERENCE_CONTEXT_PROMPT.format(reference_content=reference_context)

    # Call Claude API with vision
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompt
                    }
                ],
            }
        ],
    )

    # Parse response
    response_text = message.content[0].text

    # Try to extract JSON from response
    try:
        # Find JSON in response
        start = response_text.find('{')
        end = response_text.rfind('}') + 1
        if start != -1 and end > start:
            json_str = response_text[start:end]
            return json.loads(json_str)
    except json.JSONDecodeError:
        pass

    # Fallback: create basic metadata from filename
    clean_name = Path(filename).stem.replace("-", " ").replace("_", " ").title()
    return {
        "title": clean_name,
        "headline": f"Learn about {clean_name} with GroundSwell℠",
        "description": f"{clean_name} is a key concept in the GroundSwell℠ Business Ownership Platform.",
        "keywords": [clean_name, "GroundSwell", "business", "organizational development"],
        "category": "Business Ownership Platform",
        "supplemental_category": "General",
        "create_date": datetime.now().strftime("%Y-%m-%d")
    }


def create_xmp_packet(metadata: dict) -> str:
    """Create XMP metadata packet as XML string."""
    title = metadata.get('xmp_title', '')
    description = metadata.get('xmp_description', '')
    creator = metadata.get('xmp_creator', 'GroundSwell')
    rights = metadata.get('xmp_rights', '')
    subject = metadata.get('xmp_subject', '')
    headline = metadata.get('xmp_headline', '')
    credit = metadata.get('xmp_credit', 'GroundSwell')
    source = metadata.get('xmp_source', 'GroundSwell')
    date_created = metadata.get('xmp_date_created', '')
    category = metadata.get('xmp_category', 'Business Ownership Platform')

    # Build keywords list for XMP
    keywords = [kw.strip() for kw in subject.split(',') if kw.strip()]
    keywords_xml = '\n'.join([f'                        <rdf:li>{kw}</rdf:li>' for kw in keywords])

    xmp = f'''<?xpacket begin="" id="W5M0MpCehiHzreSzNTczkc9d"?>
<x:xmpmeta xmlns:x="adobe:ns:meta/" x:xmptk="GroundSwell Metadata Tool">
    <rdf:RDF xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#">
        <rdf:Description rdf:about=""
            xmlns:dc="http://purl.org/dc/elements/1.1/"
            xmlns:xmp="http://ns.adobe.com/xap/1.0/"
            xmlns:xmpRights="http://ns.adobe.com/xap/1.0/rights/"
            xmlns:photoshop="http://ns.adobe.com/photoshop/1.0/"
            xmlns:Iptc4xmpCore="http://iptc.org/std/Iptc4xmpCore/1.0/xmlns/"
            xmlns:iptcExt="http://iptc.org/std/Iptc4xmpExt/2008-02-29/">

            <dc:title>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{title}</rdf:li>
                </rdf:Alt>
            </dc:title>

            <dc:description>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{description}</rdf:li>
                </rdf:Alt>
            </dc:description>

            <dc:creator>
                <rdf:Seq>
                    <rdf:li>{creator}</rdf:li>
                </rdf:Seq>
            </dc:creator>

            <dc:rights>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{rights}</rdf:li>
                </rdf:Alt>
            </dc:rights>

            <dc:subject>
                <rdf:Bag>
{keywords_xml}
                </rdf:Bag>
            </dc:subject>

            <photoshop:Headline>{headline}</photoshop:Headline>
            <photoshop:Credit>{credit}</photoshop:Credit>
            <photoshop:Source>{source}</photoshop:Source>
            <photoshop:DateCreated>{date_created}</photoshop:DateCreated>
            <photoshop:Category>{category}</photoshop:Category>

            <xmpRights:Marked>True</xmpRights:Marked>

            <Iptc4xmpCore:CreatorContactInfo>
                <rdf:Description>
                    <Iptc4xmpCore:CiUrlWork>{metadata.get('xmp_website', 'www.groundswell.co')}</Iptc4xmpCore:CiUrlWork>
                    <Iptc4xmpCore:CiTelWork>{metadata.get('xmp_phone', '435-214-2997')}</Iptc4xmpCore:CiTelWork>
                </rdf:Description>
            </Iptc4xmpCore:CreatorContactInfo>

        </rdf:Description>
    </rdf:RDF>
</x:xmpmeta>
<?xpacket end="w"?>'''

    return xmp


def write_metadata_to_jpeg(file_path: Path, output_path: Path, metadata: dict) -> bool:
    """Write metadata to JPEG using piexif."""
    try:
        # Open the image
        img = Image.open(file_path)

        # Try to get existing EXIF data
        try:
            exif_dict = piexif.load(img.info.get('exif', b''))
        except:
            exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}

        # Set EXIF fields (0th IFD)
        artist = metadata.get('exif_artist', 'GroundSwell')
        copyright_text = metadata.get('exif_copyright', '')
        description = metadata.get('exif_description', '')

        exif_dict['0th'][piexif.ImageIFD.Artist] = artist.encode('utf-8')
        exif_dict['0th'][piexif.ImageIFD.Copyright] = copyright_text.encode('utf-8')
        exif_dict['0th'][piexif.ImageIFD.ImageDescription] = description.encode('utf-8')

        # Set EXIF fields (Exif IFD)
        user_comment = metadata.get('exif_user_comment', '')
        # User comment needs special encoding
        user_comment_bytes = b'ASCII\x00\x00\x00' + user_comment.encode('utf-8')
        exif_dict['Exif'][piexif.ExifIFD.UserComment] = user_comment_bytes

        # Create XMP data
        xmp_data = create_xmp_packet(metadata)

        # Dump EXIF to bytes
        exif_bytes = piexif.dump(exif_dict)

        # Save the image with EXIF
        img.save(output_path, 'JPEG', exif=exif_bytes, quality=95)

        # Now embed XMP data into the JPEG
        embed_xmp_in_jpeg(output_path, xmp_data)

        return True

    except Exception as e:
        print(f"Error writing JPEG metadata: {e}")
        import traceback
        traceback.print_exc()
        return False


def embed_xmp_in_jpeg(file_path: Path, xmp_data: str):
    """Embed XMP data into a JPEG file."""
    try:
        # Read the file
        with open(file_path, 'rb') as f:
            data = f.read()

        # Check if it's a valid JPEG
        if data[:2] != b'\xff\xd8':
            return False

        # XMP marker
        xmp_marker = b'\xff\xe1'
        xmp_header = b'http://ns.adobe.com/xap/1.0/\x00'
        xmp_bytes = xmp_data.encode('utf-8')

        # Calculate the segment length (2 bytes for length + header + data)
        segment_length = 2 + len(xmp_header) + len(xmp_bytes)
        length_bytes = segment_length.to_bytes(2, 'big')

        # Build the XMP segment
        xmp_segment = xmp_marker + length_bytes + xmp_header + xmp_bytes

        # Find position after SOI marker (first 2 bytes)
        # Insert XMP segment right after SOI
        new_data = data[:2] + xmp_segment + data[2:]

        # Write back
        with open(file_path, 'wb') as f:
            f.write(new_data)

        return True
    except Exception as e:
        print(f"Error embedding XMP: {e}")
        return False


def write_metadata_to_png(file_path: Path, output_path: Path, metadata: dict) -> bool:
    """Write metadata to PNG using Pillow's PngInfo."""
    try:
        img = Image.open(file_path)

        # Create PNG metadata
        pnginfo = PngInfo()

        # Standard PNG text chunks
        pnginfo.add_text("Title", metadata.get('xmp_title', ''))
        pnginfo.add_text("Description", metadata.get('xmp_description', ''))
        pnginfo.add_text("Author", metadata.get('exif_artist', 'GroundSwell'))
        pnginfo.add_text("Copyright", metadata.get('exif_copyright', ''))
        pnginfo.add_text("Comment", metadata.get('exif_user_comment', ''))
        pnginfo.add_text("Keywords", metadata.get('iptc_keywords', ''))
        pnginfo.add_text("Headline", metadata.get('iptc_headline', ''))
        pnginfo.add_text("Credit", metadata.get('iptc_credit', 'GroundSwell'))
        pnginfo.add_text("Source", metadata.get('xmp_source', 'GroundSwell'))
        pnginfo.add_text("Creation Time", metadata.get('exif_create_date', ''))

        # Add XMP as iTXt chunk
        xmp_data = create_xmp_packet(metadata)
        pnginfo.add_text("XML:com.adobe.xmp", xmp_data)

        # Save with metadata
        img.save(output_path, 'PNG', pnginfo=pnginfo)

        return True

    except Exception as e:
        print(f"Error writing PNG metadata: {e}")
        import traceback
        traceback.print_exc()
        return False


def write_metadata_to_image(file_path: Path, output_path: Path, metadata: dict) -> bool:
    """Write metadata to image based on file type."""
    ext = file_path.suffix.lower()

    if ext in ['.jpg', '.jpeg']:
        return write_metadata_to_jpeg(file_path, output_path, metadata)
    elif ext == '.png':
        return write_metadata_to_png(file_path, output_path, metadata)
    else:
        # For other formats, just copy the file
        shutil.copy2(file_path, output_path)
        return False


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Render the home page."""
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
async def upload_image(
    file: UploadFile = File(...),
    social_file: Optional[UploadFile] = File(None),
    featured_file: Optional[UploadFile] = File(None),
    thumbnail_file: Optional[UploadFile] = File(None),
    context_url: Optional[str] = Form(None),
    context_file: Optional[UploadFile] = File(None),
):
    """Upload and process an image with AI-generated metadata."""

    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an image (JPG, PNG, GIF, WEBP).")

    # Generate unique ID for this upload
    file_id = str(uuid.uuid4())[:8]
    original_filename = file.filename or "image.jpg"
    ext = Path(original_filename).suffix or ".jpg"

    # Save uploaded file (keep it for later editing)
    upload_path = UPLOAD_DIR / f"{file_id}_original{ext}"
    with open(upload_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Save optional variant files
    if social_file and social_file.filename and social_file.size and social_file.size > 0:
        social_ext = Path(social_file.filename).suffix or ext
        social_path = UPLOAD_DIR / f"{file_id}_social{social_ext}"
        with open(social_path, "wb") as f:
            f.write(await social_file.read())

    if featured_file and featured_file.filename and featured_file.size and featured_file.size > 0:
        featured_ext = Path(featured_file.filename).suffix or ext
        featured_path = UPLOAD_DIR / f"{file_id}_featured{featured_ext}"
        with open(featured_path, "wb") as f:
            f.write(await featured_file.read())

    if thumbnail_file and thumbnail_file.filename and thumbnail_file.size and thumbnail_file.size > 0:
        thumbnail_ext = Path(thumbnail_file.filename).suffix or ext
        thumbnail_path = UPLOAD_DIR / f"{file_id}_thumbnail{thumbnail_ext}"
        with open(thumbnail_path, "wb") as f:
            f.write(await thumbnail_file.read())

    # Get reference context if provided
    reference_context = ""
    if context_url and context_url.strip():
        reference_context = fetch_url_content(context_url.strip())
    elif context_file and context_file.filename and context_file.size and context_file.size > 0:
        file_data = await context_file.read()
        reference_context = extract_file_text(file_data, context_file.filename)

    try:
        # Analyze image with Claude
        metadata = await analyze_image_with_claude(upload_path, original_filename, reference_context)

        # Build response with file_id for later saves
        year = metadata.get("create_date", datetime.now().strftime("%Y-%m-%d"))[:4]

        return JSONResponse({
            "success": True,
            "file_id": file_id,
            "metadata": {
                "title": metadata.get("title", ""),
                "headline": metadata.get("headline", ""),
                "description": metadata.get("description", ""),
                "keywords": metadata.get("keywords", []),
                "category": metadata.get("category", "Business Ownership Platform"),
                "supplemental_category": metadata.get("supplemental_category", ""),
                "create_date": metadata.get("create_date", datetime.now().strftime("%Y-%m-%d")),
                "artist": "GroundSwell",
                "copyright": f"Copyright {year} GroundSwell",
                "credit": "GroundSwell",
                "contact": "groundswell.co",
                "website": "www.groundswell.co",
                "phone": "435-214-2997"
            }
        })

    except Exception as e:
        # Clean up on error
        upload_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"Error processing image: {str(e)}")


@app.post("/save-metadata")
async def save_metadata(request: SaveMetadataRequest):
    """Save metadata to all uploaded image variants and provide download."""

    file_id = request.file_id
    metadata = request.metadata

    upload_files = list(UPLOAD_DIR.glob(f"{file_id}_original.*"))
    if not upload_files:
        raise HTTPException(status_code=404, detail="Original file not found. Please upload again.")

    upload_path = upload_files[0]
    content_ext = upload_path.suffix
    date_str = metadata.get('exif_create_date', datetime.now().strftime('%Y-%m-%d'))

    # Build list of variants: (variant_name, source_path, file_ext)
    variants_to_process = [("content", upload_path, content_ext)]

    social_files = list(UPLOAD_DIR.glob(f"{file_id}_social.*"))
    if social_files:
        variants_to_process.append(("social", social_files[0], social_files[0].suffix))

    featured_files = list(UPLOAD_DIR.glob(f"{file_id}_featured.*"))
    if featured_files:
        variants_to_process.append(("featured", featured_files[0], featured_files[0].suffix))

    thumbnail_files = list(UPLOAD_DIR.glob(f"{file_id}_thumbnail.*"))
    if thumbnail_files:
        variants_to_process.append(("thumbnail", thumbnail_files[0], thumbnail_files[0].suffix))

    processed_files = []
    for variant_name, src_path, ext in variants_to_process:
        with open(src_path, 'rb') as f:
            img_data = f.read()

        dims = VARIANT_DIMENSIONS.get(variant_name)
        if dims:
            img_data = resize_image_to_fit(img_data, dims[0], dims[1], ext)

        temp_input = PROCESSED_DIR / f"{file_id}_temp_in{ext}"
        with open(temp_input, 'wb') as f:
            f.write(img_data)

        variant_fn = Path(upload_path.name).stem.replace(f"{file_id}_original", "image") + ext.lower()
        dl_filename = generate_download_filename(variant_fn, variant_name, date_str)
        output_path = PROCESSED_DIR / dl_filename
        write_metadata_to_image(temp_input, output_path, metadata)
        temp_input.unlink(missing_ok=True)
        processed_files.append(dl_filename)

    cleaned_stem = re.sub(r'[^a-z0-9]+', '-', Path(upload_path.name).stem.lower()).strip('-')
    zip_name = f"{date_str.replace('-', '.')}-{cleaned_stem}-images.zip"

    return JSONResponse({
        "success": True,
        "files": [
            {"filename": fname, "download_url": f"/download/{fname}", "variant": vname}
            for fname, (vname, _, _) in zip(processed_files, variants_to_process)
        ],
        "zip_name": zip_name
    })


@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download a processed image."""
    file_path = PROCESSED_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/octet-stream"
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
