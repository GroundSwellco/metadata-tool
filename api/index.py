import os
import base64
import json
import uuid
import re
import tempfile
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from io import BytesIO

from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException
from fastapi.responses import HTMLResponse, Response, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from PIL import Image
from PIL.PngImagePlugin import PngInfo
import piexif
import anthropic
import urllib.request
import urllib.error
from bs4 import BeautifulSoup
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

app = FastAPI(title="GroundSwell℠ Image Metadata Tool")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Anthropic client
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

# In-memory storage for uploaded files (temporary for serverless)
file_storage = {}

# GroundSwell metadata template
GROUNDSWELL_CONTEXT = """
You are a metadata specialist for GroundSwell℠, a Business Ownership Platform company.

GroundSwell℠ focuses on:
- Organizational Alpha (combined incremental returns and decreased risks)
- Fractal Flywheel of Organizational Development℠
- Fractal Flywheel of Management & Leadership Development℠
- Changing Pains℠ (growing pains when organization's systems don't support its size)
- Stages of Organizational Development℠
- Business Ownership School℠
- Inner Optimization℠
- Pyramid of Tech Stack Development℠
- Direct Ownership


Common keywords include: equity partners, strategic partners, return on investment, business building partner,
investment partner, Groundswell, organizational strategies, business growth, risk management, performance enhancement,
business efficiency, strategic management, value creation, operational efficiency, fractal flywheel, flywheel,
business improvement, business advice, business partner, business valuation, private equity

Categories: Business Ownership Platform, Direct Ownership, Organizational Development, Management & Leadership Development

Contact info:
- Website: www.groundswell.co
- Phone: 435-214-2997
- Credit: GroundSwell
- Copyright Status: Protected
"""

METADATA_PROMPT = """
Based on the image filename and visual content, generate metadata for this GroundSwell℠ image.

Filename: {filename}

Generate the following metadata in JSON format:
{{
    "title": "The title/name of the concept shown (from filename or image)",
    "headline": "A brief compelling headline (1 sentence) describing the value proposition",
    "description": "A detailed 2-3 sentence description of the concept",
    "keywords": ["list", "of", "relevant", "keywords", "comma", "separated"],
    "category": "Business Ownership Platform",
    "supplemental_category": "The specific sub-category (e.g., Organizational Development, Management & Leadership Development, Direct Ownership, etc.)",
    "create_date": "{current_date}"
}}

Important:
- The title should match the concept name from the filename (remove file extension, clean up formatting)
- The headline should convey the business value
- The description should explain what the concept means and its importance to businesses
- Keywords should include: the concept name, GroundSwell, relevant business terms
- Base your response on the visual content and the GroundSwell context provided

{context}

Return ONLY the JSON object, no additional text.
"""

REFERENCE_CONTEXT_PROMPT = """

Additional Context Reference:
The following content was provided as additional context for this image. Use it to generate more accurate, detailed, and contextually relevant metadata:

---
{reference_content}
---

Incorporate relevant information from this reference into the title, headline, description, and keywords.
"""

MAX_REFERENCE_LENGTH = 5000


def fetch_url_content(url: str) -> str:
    """Fetch and extract text content from a URL."""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            html = response.read().decode('utf-8', errors='ignore')
        soup = BeautifulSoup(html, 'html.parser')
        for element in soup(['script', 'style', 'nav', 'footer']):
            element.decompose()
        text = soup.get_text(separator='\n', strip=True)
        return text[:MAX_REFERENCE_LENGTH]
    except Exception:
        return ""


def extract_file_text(file_data: bytes, filename: str) -> str:
    """Extract text from uploaded file based on extension."""
    ext = Path(filename).suffix.lower()
    try:
        if ext == '.pdf':
            reader = PdfReader(BytesIO(file_data))
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
                if len(text) > MAX_REFERENCE_LENGTH:
                    break
            return text[:MAX_REFERENCE_LENGTH]
        elif ext == '.docx':
            doc = DocxDocument(BytesIO(file_data))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text[:MAX_REFERENCE_LENGTH]
        elif ext in ('.xlsx', '.xls'):
            wb = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ", ".join(str(cell) for cell in row if cell is not None)
                    if row_text:
                        text += row_text + "\n"
                    if len(text) > MAX_REFERENCE_LENGTH:
                        break
                if len(text) > MAX_REFERENCE_LENGTH:
                    break
            wb.close()
            return text[:MAX_REFERENCE_LENGTH]
        elif ext in ('.txt', '.csv', '.md', '.json'):
            text = file_data.decode('utf-8', errors='ignore')
            return text[:MAX_REFERENCE_LENGTH]
        else:
            return ""
    except Exception:
        return ""


VARIANT_DIMENSIONS = {
    "content": None,
    "social": (1200, 720),
    "featured": (700, 400),
    "thumbnail": (232, 245),
}


def resize_image_to_fit(image_data: bytes, max_width: int, max_height: int, file_ext: str) -> bytes:
    """Resize image to fit within max_width x max_height, preserving aspect ratio."""
    img = Image.open(BytesIO(image_data))
    if img.width <= max_width and img.height <= max_height:
        return image_data
    img.thumbnail((max_width, max_height), Image.LANCZOS)
    output = BytesIO()
    if file_ext.lower() in ['.jpg', '.jpeg']:
        img.save(output, 'JPEG', quality=95)
    else:
        img.save(output, 'PNG')
    return output.getvalue()


def generate_download_filename(original_filename: str, variant_type: str, date_str: str) -> str:
    """Generate standardized filename: {date}-{cleaned-name}-{type}-w.{ext}"""
    stem = Path(original_filename).stem
    ext = Path(original_filename).suffix.lower()
    cleaned = re.sub(r'[^a-z0-9]+', '-', stem.lower()).strip('-')
    if not date_str:
        date_str = datetime.now().strftime('%Y.%m.%d')
    date_dotted = date_str.replace('-', '.')
    return f"{date_dotted}-{cleaned}-{variant_type}-w{ext}"


class SaveMetadataRequest(BaseModel):
    file_id: str
    metadata: dict


def get_media_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    media_types = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp"
    }
    return media_types.get(ext, "image/jpeg")


async def analyze_image_with_claude(image_data: bytes, filename: str, reference_context: str = "") -> dict:
    """Use Claude to analyze the image and generate metadata."""

    base64_image = base64.standard_b64encode(image_data).decode("utf-8")
    media_type = get_media_type(filename)
    current_date = datetime.now().strftime("%Y-%m-%d")

    prompt = METADATA_PROMPT.format(
        filename=filename,
        current_date=current_date,
        context=GROUNDSWELL_CONTEXT
    )

    if reference_context:
        prompt += REFERENCE_CONTEXT_PROMPT.format(reference_content=reference_context)

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": base64_image,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompt
                    }
                ],
            }
        ],
    )

    response_text = message.content[0].text

    try:
        start = response_text.find('{')
        end = response_text.rfind('}') + 1
        if start != -1 and end > start:
            json_str = response_text[start:end]
            return json.loads(json_str)
    except json.JSONDecodeError:
        pass

    clean_name = Path(filename).stem.replace("-", " ").replace("_", " ").title()
    return {
        "title": clean_name,
        "headline": f"Learn about {clean_name} with GroundSwell℠",
        "description": f"{clean_name} is a key concept in the GroundSwell℠ Business Ownership Platform.",
        "keywords": [clean_name, "GroundSwell", "business", "organizational development"],
        "category": "Business Ownership Platform",
        "supplemental_category": "General",
        "create_date": datetime.now().strftime("%Y-%m-%d")
    }


def create_xmp_packet(metadata: dict) -> str:
    """Create XMP metadata packet as XML string."""
    title = metadata.get('xmp_title', '')
    description = metadata.get('xmp_description', '')
    creator = metadata.get('xmp_creator', 'GroundSwell')
    rights = metadata.get('xmp_rights', '')
    subject = metadata.get('xmp_subject', '')
    headline = metadata.get('xmp_headline', '')
    credit = metadata.get('xmp_credit', 'GroundSwell')
    source = metadata.get('xmp_source', 'GroundSwell')
    date_created = metadata.get('xmp_date_created', '')
    category = metadata.get('xmp_category', 'Business Ownership Platform')

    keywords = [kw.strip() for kw in subject.split(',') if kw.strip()]
    keywords_xml = '\n'.join([f'                        <rdf:li>{kw}</rdf:li>' for kw in keywords])

    xmp = f'''<?xpacket begin="" id="W5M0MpCehiHzreSzNTczkc9d"?>
<x:xmpmeta xmlns:x="adobe:ns:meta/" x:xmptk="GroundSwell Metadata Tool">
    <rdf:RDF xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#">
        <rdf:Description rdf:about=""
            xmlns:dc="http://purl.org/dc/elements/1.1/"
            xmlns:xmp="http://ns.adobe.com/xap/1.0/"
            xmlns:xmpRights="http://ns.adobe.com/xap/1.0/rights/"
            xmlns:photoshop="http://ns.adobe.com/photoshop/1.0/"
            xmlns:Iptc4xmpCore="http://iptc.org/std/Iptc4xmpCore/1.0/xmlns/">

            <dc:title>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{title}</rdf:li>
                </rdf:Alt>
            </dc:title>

            <dc:description>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{description}</rdf:li>
                </rdf:Alt>
            </dc:description>

            <dc:creator>
                <rdf:Seq>
                    <rdf:li>{creator}</rdf:li>
                </rdf:Seq>
            </dc:creator>

            <dc:rights>
                <rdf:Alt>
                    <rdf:li xml:lang="x-default">{rights}</rdf:li>
                </rdf:Alt>
            </dc:rights>

            <dc:subject>
                <rdf:Bag>
{keywords_xml}
                </rdf:Bag>
            </dc:subject>

            <photoshop:Headline>{headline}</photoshop:Headline>
            <photoshop:Credit>{credit}</photoshop:Credit>
            <photoshop:Source>{source}</photoshop:Source>
            <photoshop:DateCreated>{date_created}</photoshop:DateCreated>
            <photoshop:Category>{category}</photoshop:Category>

            <xmpRights:Marked>True</xmpRights:Marked>

            <Iptc4xmpCore:CreatorContactInfo>
                <rdf:Description>
                    <Iptc4xmpCore:CiUrlWork>{metadata.get('xmp_website', 'www.groundswell.co')}</Iptc4xmpCore:CiUrlWork>
                    <Iptc4xmpCore:CiTelWork>{metadata.get('xmp_phone', '435-214-2997')}</Iptc4xmpCore:CiTelWork>
                </rdf:Description>
            </Iptc4xmpCore:CreatorContactInfo>

        </rdf:Description>
    </rdf:RDF>
</x:xmpmeta>
<?xpacket end="w"?>'''

    return xmp


def process_jpeg_metadata(image_data: bytes, metadata: dict) -> bytes:
    """Process JPEG and add metadata, return bytes."""
    img = Image.open(BytesIO(image_data))

    try:
        exif_dict = piexif.load(img.info.get('exif', b''))
    except:
        exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}, "thumbnail": None}

    artist = metadata.get('exif_artist', 'GroundSwell')
    copyright_text = metadata.get('exif_copyright', '')
    description = metadata.get('exif_description', '')

    exif_dict['0th'][piexif.ImageIFD.Artist] = artist.encode('utf-8')
    exif_dict['0th'][piexif.ImageIFD.Copyright] = copyright_text.encode('utf-8')
    exif_dict['0th'][piexif.ImageIFD.ImageDescription] = description.encode('utf-8')

    user_comment = metadata.get('exif_user_comment', '')
    user_comment_bytes = b'ASCII\x00\x00\x00' + user_comment.encode('utf-8')
    exif_dict['Exif'][piexif.ExifIFD.UserComment] = user_comment_bytes

    exif_bytes = piexif.dump(exif_dict)

    output = BytesIO()
    img.save(output, 'JPEG', exif=exif_bytes, quality=95)
    jpeg_data = output.getvalue()

    # Embed XMP
    xmp_data = create_xmp_packet(metadata)
    xmp_marker = b'\xff\xe1'
    xmp_header = b'http://ns.adobe.com/xap/1.0/\x00'
    xmp_bytes = xmp_data.encode('utf-8')

    segment_length = 2 + len(xmp_header) + len(xmp_bytes)
    length_bytes = segment_length.to_bytes(2, 'big')

    xmp_segment = xmp_marker + length_bytes + xmp_header + xmp_bytes
    final_data = jpeg_data[:2] + xmp_segment + jpeg_data[2:]

    return final_data


def process_png_metadata(image_data: bytes, metadata: dict) -> bytes:
    """Process PNG and add metadata, return bytes."""
    img = Image.open(BytesIO(image_data))

    pnginfo = PngInfo()
    pnginfo.add_text("Title", metadata.get('xmp_title', ''))
    pnginfo.add_text("Description", metadata.get('xmp_description', ''))
    pnginfo.add_text("Author", metadata.get('exif_artist', 'GroundSwell'))
    pnginfo.add_text("Copyright", metadata.get('exif_copyright', ''))
    pnginfo.add_text("Comment", metadata.get('exif_user_comment', ''))
    pnginfo.add_text("Keywords", metadata.get('iptc_keywords', ''))
    pnginfo.add_text("Headline", metadata.get('iptc_headline', ''))
    pnginfo.add_text("Credit", metadata.get('iptc_credit', 'GroundSwell'))
    pnginfo.add_text("Source", metadata.get('xmp_source', 'GroundSwell'))
    pnginfo.add_text("Creation Time", metadata.get('exif_create_date', ''))

    xmp_data = create_xmp_packet(metadata)
    pnginfo.add_text("XML:com.adobe.xmp", xmp_data)

    output = BytesIO()
    img.save(output, 'PNG', pnginfo=pnginfo)
    return output.getvalue()


def process_image_metadata(image_data: bytes, filename: str, metadata: dict) -> bytes:
    """Process image and add metadata based on file type."""
    ext = Path(filename).suffix.lower()

    if ext in ['.jpg', '.jpeg']:
        return process_jpeg_metadata(image_data, metadata)
    elif ext == '.png':
        return process_png_metadata(image_data, metadata)
    else:
        return image_data


# HTML Template embedded
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GroundSwell Image Metadata Tool</title>
    <style>
        @font-face { font-family: 'NHaasGroteskTXPro'; src: url('/fonts/NHaasGroteskTXPro-65Md.ttf') format('truetype'); font-weight: 500; font-style: normal; font-display: swap; }
        @font-face { font-family: 'NeuzeitGro-Reg'; src: url('/fonts/NeuzeitGro-Reg.ttf') format('truetype'); font-weight: 400; font-style: normal; font-display: swap; }
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'NeuzeitGro-Reg', 'Helvetica Neue', Helvetica, Arial, sans-serif; background: linear-gradient(135deg, #2C3B4C 0%, #1F2A36 100%); min-height: 100vh; color: #fff; }
        h1, h2, h3, h4, h5, h6, .logo, .footer-brand .logo-small, .btn, .btn-primary, .btn-secondary, .new-upload-btn { font-family: 'NHaasGroteskTXPro', 'Helvetica Neue', Helvetica, Arial, sans-serif; }
        .container { max-width: 1200px; margin: 0 auto; padding: 40px 20px; }
        header { text-align: center; margin-bottom: 40px; }
        .logo { font-size: 2.5rem; font-weight: 700; color: #fff; }
        .subtitle { color: #94a3b8; margin-top: 8px; }
        .upload-section { background: rgba(255,255,255,0.05); border: 2px dashed rgba(255,255,255,0.2); border-radius: 16px; padding: 60px 40px; text-align: center; cursor: pointer; transition: all 0.3s; }
        .upload-section:hover { border-color: #7FBBE6; background: rgba(127,187,230,0.05); }
        .upload-icon { font-size: 4rem; margin-bottom: 20px; }
        #fileInput { display: none; }
        .processing { display: none; text-align: center; padding: 40px; }
        .spinner { width: 60px; height: 60px; border: 4px solid rgba(255,255,255,0.1); border-left-color: #7FBBE6; border-radius: 50%; animation: spin 1s linear infinite; margin: 0 auto 20px; }
        @keyframes spin { to { transform: rotate(360deg); } }
        .results { display: none; margin-top: 40px; }
        .result-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 30px; padding-bottom: 20px; border-bottom: 1px solid rgba(255,255,255,0.1); }
        .result-title { font-size: 1.8rem; color: #fff; }
        .btn { color: #fff; border: none; padding: 12px 24px; border-radius: 8px; font-size: 1rem; cursor: pointer; transition: all 0.2s; }
        .btn-primary { background: #F16365; color: #fff; box-shadow: 0 4px 14px rgba(241,99,101,0.35); }
        .btn-primary:hover { background: #ef4f51; box-shadow: 0 6px 18px rgba(241,99,101,0.5); }
        .btn-secondary { background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.2); }
        .btn:hover { transform: translateY(-2px); box-shadow: 0 4px 20px rgba(127,187,230,0.3); }
        .metadata-tabs { display: flex; gap: 4px; margin-bottom: 20px; background: rgba(0,0,0,0.2); padding: 4px; border-radius: 12px; }
        .tab-btn { flex: 1; padding: 12px 20px; background: transparent; border: none; color: #94a3b8; cursor: pointer; border-radius: 8px; transition: all 0.3s; }
        .tab-btn.active { background: rgba(241,99,101,0.3); color: #fff; }
        .metadata-panel { display: none; background: rgba(255,255,255,0.05); border-radius: 16px; padding: 30px; }
        .metadata-panel.active { display: block; }
        .metadata-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 16px; }
        .metadata-field { background: rgba(0,0,0,0.2); border-radius: 10px; padding: 16px; }
        .metadata-field.full-width { grid-column: 1 / -1; }
        .field-label { font-size: 0.75rem; color: #64748b; text-transform: uppercase; margin-bottom: 8px; display: flex; justify-content: space-between; }
        .field-tag { font-size: 0.65rem; padding: 2px 6px; border-radius: 4px; }
        .tag-exif { background: rgba(239,68,68,0.2); color: #fca5a5; }
        .tag-iptc { background: rgba(34,197,94,0.2); color: #86efac; }
        .tag-xmp { background: rgba(59,130,246,0.2); color: #93c5fd; }
        .field-input { width: 100%; background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); border-radius: 6px; padding: 10px 12px; color: #e2e8f0; font-size: 0.95rem; }
        .field-input:focus { outline: none; border-color: #F16365; }
        textarea.field-input { min-height: 80px; resize: vertical; }
        .status-message { padding: 12px 16px; border-radius: 8px; margin-bottom: 20px; display: none; }
        .status-message.success { background: rgba(34,197,94,0.1); border: 1px solid rgba(34,197,94,0.3); color: #86efac; display: block; }
        .status-message.error { background: rgba(239,68,68,0.1); border: 1px solid rgba(239,68,68,0.3); color: #fca5a5; display: block; }
        .new-upload-btn { display: block; width: 100%; padding: 16px; background: transparent; border: 2px solid rgba(255,255,255,0.2); color: #94a3b8; border-radius: 8px; cursor: pointer; margin-top: 20px; }
        .new-upload-btn:hover { border-color: #7FBBE6; color: #7FBBE6; }
        footer { text-align: center; margin-top: 60px; color: #64748b; }
        footer a { color: #7FBBE6; text-decoration: none; }
        .saving-overlay { position: fixed; inset: 0; background: rgba(0,0,0,0.7); display: none; justify-content: center; align-items: center; z-index: 1000; }
        .saving-overlay.active { display: flex; }
        .button-group { display: flex; gap: 12px; }
        .panel-header { margin-bottom: 20px; }
        .panel-header h3 { color: #fff; font-size: 1.1rem; margin-bottom: 4px; }
        .panel-header p { color: #64748b; font-size: 0.85rem; }
        .tab-badge { display: inline-block; background: rgba(127,187,230,0.2); color: #7FBBE6; padding: 2px 8px; border-radius: 10px; font-size: 0.75rem; margin-left: 8px; }
        .preview-section { display: none; }
        .preview-card { background: rgba(255,255,255,0.05); border-radius: 16px; padding: 30px; text-align: center; margin-bottom: 20px; }
        .preview-card img { max-height: 200px; max-width: 100%; border-radius: 8px; margin-bottom: 16px; display: block; margin-left: auto; margin-right: auto; }
        .preview-filename { color: #94a3b8; font-size: 0.95rem; margin-bottom: 12px; }
        .generate-btn { display: block; width: 100%; padding: 16px; font-size: 1.1rem; margin-top: 20px; }
        .variant-uploads-section { margin-bottom: 20px; }
        .variant-uploads-section h3 { color: #fff; font-size: 1.1rem; margin-bottom: 4px; }
        .variant-uploads-section > p { color: #64748b; font-size: 0.85rem; margin-bottom: 16px; }
        .variant-upload-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
        .variant-upload-card { background: rgba(255,255,255,0.05); border-radius: 12px; padding: 16px; }
        .variant-upload-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; color: #e2e8f0; font-weight: 500; }
        .variant-size { color: #64748b; font-size: 0.8rem; }
        .variant-upload-zone { border: 2px dashed rgba(255,255,255,0.15); border-radius: 8px; padding: 20px 16px; cursor: pointer; transition: all 0.3s; min-height: 80px; display: flex; flex-direction: column; align-items: center; justify-content: center; text-align: center; }
        .variant-upload-zone:hover { border-color: #7FBBE6; background: rgba(127,187,230,0.05); }
        .variant-upload-zone.has-file { border-color: rgba(241,99,101,0.4); border-style: solid; }
        .variant-upload-zone img { max-height: 80px; max-width: 100%; border-radius: 4px; margin-bottom: 8px; }
        .variant-upload-zone p { color: #64748b; font-size: 0.85rem; margin: 0; }
        .variant-remove-btn { margin-top: 8px; font-size: 0.8rem; padding: 4px 12px; }
        .download-section { background: rgba(255,255,255,0.05); border-radius: 12px; padding: 20px; margin-bottom: 20px; display: none; }
        .download-section h3 { color: #fff; font-size: 1.1rem; margin-bottom: 16px; }
        .download-item { display: flex; justify-content: space-between; align-items: center; padding: 12px 16px; background: rgba(0,0,0,0.2); border-radius: 8px; margin-bottom: 8px; }
        .download-item-name { color: #e2e8f0; font-size: 0.9rem; word-break: break-all; margin-right: 12px; }
        .download-item-type { color: #64748b; font-size: 0.75rem; text-transform: uppercase; margin-right: auto; padding-left: 8px; }
        .download-all-row { margin-top: 16px; text-align: center; }
        .variant-select-section { margin-top: 16px; }
        .variant-select-header { cursor: pointer; color: #94a3b8; display: flex; justify-content: center; gap: 8px; align-items: center; padding: 12px; transition: color 0.3s; font-size: 0.95rem; }
        .variant-select-header:hover { color: #7FBBE6; }
        .variant-select-body { display: none; padding-top: 12px; }
        .variant-select-body.open { display: block; }
        .variant-select-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; }
        .variant-select-card { background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; padding: 16px; text-align: center; cursor: pointer; transition: all 0.3s; }
        .variant-select-card:hover { border-color: #7FBBE6; background: rgba(127,187,230,0.05); }
        .variant-select-card .vs-name { display: block; color: #e2e8f0; font-weight: 500; margin-bottom: 4px; }
        .variant-select-card .vs-size { display: block; color: #64748b; font-size: 0.8rem; }
        .context-section { margin-top: 20px; background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.1); border-radius: 12px; overflow: hidden; }
        .context-header { padding: 16px 20px; cursor: pointer; display: flex; justify-content: space-between; align-items: center; color: #94a3b8; transition: color 0.3s; user-select: none; }
        .context-header:hover { color: #7FBBE6; }
        .context-body { padding: 0 20px 20px; display: none; }
        .context-body.open { display: block; }
        .context-field { margin-bottom: 12px; }
        .context-field .field-label { font-size: 0.8rem; color: #94a3b8; margin-bottom: 6px; text-transform: uppercase; letter-spacing: 0.5px; }
        .context-divider { text-align: center; color: #475569; margin: 16px 0; font-size: 0.85rem; }
        .pdf-upload { display: flex; align-items: center; gap: 12px; flex-wrap: wrap; }
        .btn-small { padding: 8px 16px; font-size: 0.85rem; }
        .context-badge { display: inline-block; background: rgba(34,197,94,0.2); color: #86efac; padding: 2px 8px; border-radius: 10px; font-size: 0.75rem; margin-left: 8px; display: none; }
        .hero { text-align: center; padding: 60px 0 30px; }
        .hero .logo { font-size: 1.4rem; margin-bottom: 20px; }
        .hero h1 { font-size: 3rem; font-weight: 800; line-height: 1.15; margin-bottom: 20px; color: #fff; }
        .hero-lead { font-size: 1.2rem; color: #94a3b8; max-width: 720px; margin: 0 auto 24px; line-height: 1.5; }
        .hero-pills { display: flex; justify-content: center; gap: 10px; flex-wrap: wrap; margin-top: 24px; }
        .hero-pill { background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); border-radius: 999px; padding: 8px 16px; font-size: 0.9rem; color: #cbd5e1; }
        .marketing { margin-top: 80px; }
        .marketing-section { margin-bottom: 80px; }
        .marketing-section h2 { font-size: 2rem; text-align: center; margin-bottom: 12px; color: #fff; }
        .section-lead { text-align: center; color: #94a3b8; max-width: 640px; margin: 0 auto 40px; font-size: 1.05rem; line-height: 1.5; }
        .steps-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 24px; }
        .step-card { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08); border-radius: 16px; padding: 32px 28px; transition: transform 0.2s ease, border-color 0.2s ease; }
        .step-card:hover { transform: translateY(-4px); border-color: rgba(127,187,230,0.4); }
        .step-number { font-size: 0.8rem; font-weight: 700; color: #7FBBE6; letter-spacing: 2px; margin-bottom: 12px; }
        .step-card h3 { font-size: 1.3rem; margin-bottom: 10px; }
        .step-card p { color: #94a3b8; line-height: 1.6; font-size: 0.95rem; }
        .features-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; }
        .feature-card { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08); border-radius: 14px; padding: 24px; transition: transform 0.2s ease, border-color 0.2s ease; }
        .feature-card:hover { transform: translateY(-3px); border-color: rgba(241,99,101,0.4); }
        .feature-icon { font-size: 1.8rem; margin-bottom: 12px; }
        .feature-card h4 { font-size: 1.1rem; margin-bottom: 8px; }
        .feature-card p { color: #94a3b8; font-size: 0.9rem; line-height: 1.5; }
        .brand-defaults { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08); border-radius: 16px; padding: 32px 40px; }
        .defaults-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 0 48px; }
        .default-row { display: flex; justify-content: space-between; gap: 16px; padding: 12px 0; border-bottom: 1px solid rgba(255,255,255,0.06); }
        .default-key { color: #94a3b8; font-size: 0.9rem; }
        .default-value { color: #fff; font-weight: 500; font-size: 0.95rem; text-align: right; }
        .cta-band { background: linear-gradient(135deg, rgba(127,187,230,0.08), rgba(241,99,101,0.08)); border: 1px solid rgba(127,187,230,0.2); border-radius: 20px; padding: 48px 40px; text-align: center; }
        .cta-band h2 { font-size: 1.8rem; margin-bottom: 12px; -webkit-text-fill-color: #fff; background: none; color: #fff; }
        .cta-band p { color: #94a3b8; margin-bottom: 24px; }
        .brand-footer { margin-top: 80px; padding: 40px 0 20px; border-top: 1px solid rgba(255,255,255,0.08); }
        .footer-grid { display: grid; grid-template-columns: 2fr 1fr 1fr; gap: 40px; margin-bottom: 24px; }
        .footer-brand .logo-small { font-size: 1.4rem; font-weight: 700; color: #fff; margin-bottom: 10px; }
        .footer-brand p { color: #94a3b8; font-size: 0.9rem; line-height: 1.5; max-width: 340px; }
        .footer-col h5 { color: #cbd5e1; margin-bottom: 12px; font-size: 0.95rem; }
        .footer-col a { color: #94a3b8; text-decoration: none; font-size: 0.9rem; display: block; padding: 4px 0; }
        .footer-col a:hover { color: #7FBBE6; }
        .footer-bottom { text-align: center; color: #64748b; font-size: 0.85rem; padding-top: 20px; border-top: 1px solid rgba(255,255,255,0.05); }
        .footer-bottom a { color: #94a3b8; text-decoration: none; }
        .footer-bottom a:hover { color: #7FBBE6; }
        .pricing-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 32px; }
        .plan-card { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.08); border-radius: 16px; padding: 32px 28px; display: flex; flex-direction: column; transition: transform 0.2s ease, border-color 0.2s ease; position: relative; }
        .plan-card:hover { transform: translateY(-4px); border-color: rgba(127,187,230,0.4); }
        .plan-card.featured { border-color: rgba(241,99,101,0.5); background: linear-gradient(180deg, rgba(241,99,101,0.10), rgba(127,187,230,0.04)); }
        .plan-badge { position: absolute; top: -12px; right: 24px; background: linear-gradient(90deg, #7FBBE6, #F16365); padding: 4px 12px; border-radius: 999px; font-size: 0.72rem; font-weight: 700; letter-spacing: 0.5px; }
        .plan-name { color: #cbd5e1; text-transform: uppercase; font-size: 0.8rem; letter-spacing: 2px; margin-bottom: 12px; }
        .plan-price { font-size: 2.5rem; font-weight: 800; margin-bottom: 4px; }
        .plan-price .per { font-size: 0.9rem; color: #94a3b8; font-weight: 400; }
        .plan-desc { color: #94a3b8; margin-bottom: 20px; font-size: 0.95rem; }
        .plan-features { list-style: none; padding: 0; margin: 0 0 24px; flex-grow: 1; }
        .plan-features li { color: #cbd5e1; font-size: 0.9rem; padding: 6px 0 6px 22px; position: relative; }
        .plan-features li:before { content: '\2713'; position: absolute; left: 0; color: #7FBBE6; font-weight: 700; }
        .plan-cta { width: 100%; }
        .credits-card { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08); border-radius: 16px; padding: 28px 36px; display: flex; flex-wrap: wrap; gap: 24px; align-items: center; justify-content: space-between; }
        .credits-info h3 { font-size: 1.3rem; margin-bottom: 6px; }
        .credits-info p { color: #94a3b8; font-size: 0.95rem; max-width: 380px; }
        .credits-packs { display: flex; gap: 10px; flex-wrap: wrap; }
        .pack-btn { background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.1); color: #e2e8f0; padding: 12px 18px; border-radius: 10px; cursor: pointer; transition: all 0.2s; text-align: center; }
        .pack-btn:hover { border-color: #7FBBE6; background: rgba(127,187,230,0.08); transform: translateY(-2px); }
        .pack-btn .pack-qty { display: block; font-weight: 700; font-size: 1.05rem; }
        .pack-btn .pack-price { display: block; color: #94a3b8; font-size: 0.78rem; margin-top: 2px; }
        .cart-fab { position: fixed; bottom: 24px; right: 24px; width: 60px; height: 60px; border-radius: 50%; background: linear-gradient(135deg, #7FBBE6, #F16365); border: none; cursor: pointer; box-shadow: 0 8px 24px rgba(241,99,101,0.4); display: flex; align-items: center; justify-content: center; font-size: 1.6rem; z-index: 500; transition: transform 0.2s; }
        .cart-fab:hover { transform: scale(1.05); }
        .cart-badge { position: absolute; top: -4px; right: -4px; background: #ef4444; color: #fff; border-radius: 999px; padding: 2px 7px; font-size: 0.72rem; font-weight: 700; min-width: 20px; display: none; }
        .cart-badge.visible { display: block; }
        .cart-backdrop { position: fixed; inset: 0; background: rgba(0,0,0,0.6); z-index: 998; opacity: 0; pointer-events: none; transition: opacity 0.3s; }
        .cart-backdrop.open { opacity: 1; pointer-events: auto; }
        .cart-drawer { position: fixed; top: 0; right: 0; width: 420px; max-width: 100vw; height: 100vh; background: #0f1729; border-left: 1px solid rgba(255,255,255,0.1); z-index: 999; transform: translateX(100%); transition: transform 0.3s ease; display: flex; flex-direction: column; }
        .cart-drawer.open { transform: translateX(0); }
        .cart-header { padding: 22px 24px; border-bottom: 1px solid rgba(255,255,255,0.08); display: flex; justify-content: space-between; align-items: center; }
        .cart-header h3 { font-size: 1.25rem; }
        .cart-close { background: none; border: none; color: #94a3b8; font-size: 1.6rem; cursor: pointer; padding: 0 8px; border-radius: 6px; line-height: 1; }
        .cart-close:hover { background: rgba(255,255,255,0.05); color: #fff; }
        .cart-body { flex: 1; overflow-y: auto; padding: 20px 24px; }
        .cart-empty { text-align: center; color: #64748b; padding: 60px 20px; }
        .cart-empty-icon { font-size: 3rem; margin-bottom: 12px; }
        .cart-item { background: rgba(255,255,255,0.04); border-radius: 12px; padding: 16px; margin-bottom: 12px; }
        .cart-item-top { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 10px; gap: 12px; }
        .cart-item-name { font-weight: 600; margin-bottom: 2px; }
        .cart-item-meta { font-size: 0.8rem; color: #64748b; }
        .cart-item-price { color: #7FBBE6; font-weight: 600; }
        .cart-item-remove { background: none; border: none; color: #64748b; cursor: pointer; font-size: 1.2rem; padding: 0 8px; border-radius: 4px; line-height: 1; }
        .cart-item-remove:hover { color: #ef4444; background: rgba(239,68,68,0.08); }
        .cart-qty { display: inline-flex; align-items: center; gap: 6px; background: rgba(0,0,0,0.25); border-radius: 8px; padding: 4px; }
        .cart-qty button { width: 28px; height: 28px; border-radius: 6px; background: rgba(255,255,255,0.06); border: none; color: #fff; cursor: pointer; font-size: 1rem; }
        .cart-qty button:hover { background: rgba(255,255,255,0.12); }
        .cart-qty span { min-width: 32px; text-align: center; font-weight: 600; }
        .cart-footer { padding: 20px 24px; border-top: 1px solid rgba(255,255,255,0.08); background: rgba(0,0,0,0.25); }
        .cart-totals { margin-bottom: 16px; }
        .cart-row { display: flex; justify-content: space-between; padding: 4px 0; color: #94a3b8; font-size: 0.9rem; }
        .cart-row.total { color: #fff; font-size: 1.15rem; font-weight: 700; padding-top: 10px; margin-top: 6px; border-top: 1px solid rgba(255,255,255,0.08); }
        .checkout-btn { width: 100%; padding: 15px; font-size: 1.05rem; font-weight: 600; }
        .checkout-btn:disabled { opacity: 0.4; cursor: not-allowed; transform: none; box-shadow: none; }
        .modal-backdrop { position: fixed; inset: 0; background: rgba(0,0,0,0.75); z-index: 1100; display: none; align-items: center; justify-content: center; padding: 20px; }
        .modal-backdrop.open { display: flex; }
        .modal { background: #0f1729; border: 1px solid rgba(241,99,101,0.3); border-radius: 16px; padding: 40px; max-width: 440px; width: 100%; text-align: center; }
        .modal-icon { font-size: 3rem; margin-bottom: 16px; }
        .modal h3 { font-size: 1.4rem; margin-bottom: 10px; }
        .modal p { color: #94a3b8; margin-bottom: 24px; line-height: 1.5; }
        @media (max-width: 768px) { .hero h1 { font-size: 2rem; } .steps-grid, .features-grid, .defaults-grid, .pricing-grid { grid-template-columns: 1fr; } .footer-grid { grid-template-columns: 1fr; } .cart-drawer { width: 100vw; } .credits-card { flex-direction: column; align-items: stretch; text-align: center; } .credits-packs { justify-content: center; } }
    </style>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
</head>
<body>
    <div class="container">
        <section class="hero">
            <div class="logo">GroundSwell℠</div>
            <h1>AI-Powered Metadata,<br>Tailored to Your Brand</h1>
            <p class="hero-lead">Drop an image and get publish-ready EXIF, IPTC, and XMP metadata crafted in the GroundSwell voice — in seconds, not hours.</p>
            <div class="hero-pills">
                <span class="hero-pill">⚡ Powered by Claude Sonnet 4</span>
                <span class="hero-pill">🔒 Brand-safe defaults</span>
                <span class="hero-pill">📦 Multi-variant export</span>
            </div>
        </section>

        <div class="upload-section" id="uploadSection">
            <div class="upload-icon">📷</div>
            <p>Drag & drop your image here or <span style="color:#7FBBE6">browse files</span></p>
            <p style="color:#64748b;margin-top:10px;font-size:0.9rem">Supports JPG, PNG</p>
            <input type="file" id="fileInput" accept="image/*">
        </div>

        <div class="variant-select-section" id="variantSelectSection">
            <p style="color:#94a3b8;text-align:center;margin-bottom:12px;font-size:0.95rem;">Or upload a specific variant</p>
            <div class="variant-select-grid">
                    <div class="variant-select-card" data-variant="social">
                        <span class="vs-name">Social</span><span class="vs-size">1200 x 720</span>
                        <input type="file" accept="image/*" style="display:none">
                    </div>
                    <div class="variant-select-card" data-variant="featured">
                        <span class="vs-name">Featured</span><span class="vs-size">700 x 400</span>
                        <input type="file" accept="image/*" style="display:none">
                    </div>
                    <div class="variant-select-card" data-variant="thumbnail">
                        <span class="vs-name">Thumbnail</span><span class="vs-size">232 x 245</span>
                        <input type="file" accept="image/*" style="display:none">
                    </div>
                </div>
        </div>

        <div class="preview-section" id="previewSection">
            <div class="preview-card">
                <img id="imagePreview" src="" alt="Preview">
                <p class="preview-filename" id="previewFilename"></p>
                <span style="color:#64748b;font-size:0.8rem;" id="previewVariantLabel">Content (full size)</span>
                <div style="margin-top:12px;"><button class="btn btn-secondary btn-small" type="button" id="changeImageBtn">Change Image</button></div>
            </div>

            <div class="variant-uploads-section">
                <h3>Additional Variants (Optional)</h3>
                <p>Upload separate image files for social and featured sizes</p>
                <div class="variant-upload-grid">
                    <div class="variant-upload-card">
                        <div class="variant-upload-header"><span>Social</span><span class="variant-size">1200 x 720</span></div>
                        <div class="variant-upload-zone" id="socialUploadZone">
                            <img id="socialPreview" src="" alt="" style="display:none">
                            <p id="socialPlaceholder">Click or drop image</p>
                            <input type="file" id="socialInput" accept="image/*" style="display:none">
                        </div>
                        <button class="btn btn-secondary variant-remove-btn" id="socialRemoveBtn" style="display:none">Remove</button>
                    </div>
                    <div class="variant-upload-card">
                        <div class="variant-upload-header"><span>Featured</span><span class="variant-size">700 x 400</span></div>
                        <div class="variant-upload-zone" id="featuredUploadZone">
                            <img id="featuredPreview" src="" alt="" style="display:none">
                            <p id="featuredPlaceholder">Click or drop image</p>
                            <input type="file" id="featuredInput" accept="image/*" style="display:none">
                        </div>
                        <button class="btn btn-secondary variant-remove-btn" id="featuredRemoveBtn" style="display:none">Remove</button>
                    </div>
                    <div class="variant-upload-card">
                        <div class="variant-upload-header"><span>Thumbnail</span><span class="variant-size">232 x 245</span></div>
                        <div class="variant-upload-zone" id="thumbnailUploadZone">
                            <img id="thumbnailPreview" src="" alt="" style="display:none">
                            <p id="thumbnailPlaceholder">Click or drop image</p>
                            <input type="file" id="thumbnailInput" accept="image/*" style="display:none">
                        </div>
                        <button class="btn btn-secondary variant-remove-btn" id="thumbnailRemoveBtn" style="display:none">Remove</button>
                    </div>
                </div>
            </div>

            <div class="context-section">
                <div class="context-header" id="contextHeader">
                    <span>Context Reference (Optional)<span class="context-badge" id="contextBadge">Added</span></span>
                    <span class="toggle-icon" id="contextToggle">&#9660;</span>
                </div>
                <div class="context-body" id="contextBody">
                    <p style="color:#64748b;font-size:0.85rem;margin-bottom:16px;">Provide a URL or file (PDF, Word, Excel, TXT) to help the AI generate more accurate metadata</p>
                    <div class="context-field">
                        <div class="field-label">Website URL</div>
                        <input type="url" class="field-input" id="contextUrl" placeholder="https://example.com/article-about-this-image">
                    </div>
                    <div class="context-divider">&#8212; or &#8212;</div>
                    <div class="context-field">
                        <div class="field-label">Upload File</div>
                        <div class="pdf-upload">
                            <button class="btn btn-secondary btn-small" type="button" id="pdfBtn">Choose File</button>
                            <span style="color:#94a3b8;font-size:0.85rem;" id="pdfFileName">No file selected</span>
                            <input type="file" id="pdfInput" accept=".pdf,.docx,.xlsx,.xls,.txt,.csv,.md,.json" style="display:none">
                        </div>
                    </div>
                </div>
            </div>

            <button class="btn btn-primary generate-btn" id="generateBtn">Generate Metadata</button>
        </div>

        <div class="processing" id="processing">
            <div class="spinner"></div>
            <p>Analyzing image with AI...</p>
        </div>

        <div class="results" id="results">
            <div class="result-header">
                <h2 class="result-title" id="resultTitle">Image Metadata</h2>
                <div class="button-group">
                    <button class="btn btn-secondary" id="resetBtn">Reset to AI Generated</button>
                    <button class="btn btn-primary" id="saveDownloadBtn">Save & Download</button>
                </div>
            </div>

            <div id="statusMessage" class="status-message"></div>

            <div class="download-section" id="downloadSection">
                <h3>Download Files</h3>
                <div id="downloadList"></div>
                <div class="download-all-row" id="downloadAllRow" style="display:none;">
                    <button class="btn btn-primary" id="downloadAllBtn">Download All as Zip</button>
                </div>
            </div>

            <div class="metadata-tabs">
                <button class="tab-btn active" data-tab="exif">EXIF <span class="tab-badge">5</span></button>
                <button class="tab-btn" data-tab="iptc">IPTC <span class="tab-badge">9</span></button>
                <button class="tab-btn" data-tab="xmp">XMP <span class="tab-badge">14</span></button>
            </div>

            <div class="metadata-panel active" id="panel-exif">
                <div class="panel-header"><h3>EXIF - Exchangeable Image File Format</h3><p>Standard metadata embedded in image files</p></div>
                <div class="metadata-grid">
                    <div class="metadata-field"><div class="field-label">Create Date <span class="field-tag tag-exif">EXIF</span></div><input type="date" class="field-input" id="exif_create_date"></div>
                    <div class="metadata-field"><div class="field-label">Artist <span class="field-tag tag-exif">EXIF</span></div><input type="text" class="field-input" id="exif_artist" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Copyright <span class="field-tag tag-exif">EXIF</span></div><input type="text" class="field-input" id="exif_copyright"></div>
                    <div class="metadata-field full-width"><div class="field-label">Image Description <span class="field-tag tag-exif">EXIF</span></div><textarea class="field-input" id="exif_description"></textarea></div>
                    <div class="metadata-field full-width"><div class="field-label">User Comment (Keywords) <span class="field-tag tag-exif">EXIF</span></div><textarea class="field-input" id="exif_user_comment"></textarea></div>
                </div>
            </div>

            <div class="metadata-panel" id="panel-iptc">
                <div class="panel-header"><h3>IPTC - International Press Telecommunications Council</h3><p>Industry standard for news and media metadata</p></div>
                <div class="metadata-grid">
                    <div class="metadata-field"><div class="field-label">Object Name/Title <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_object_name"></div>
                    <div class="metadata-field"><div class="field-label">Headline <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_headline"></div>
                    <div class="metadata-field full-width"><div class="field-label">Caption/Abstract <span class="field-tag tag-iptc">IPTC</span></div><textarea class="field-input" id="iptc_caption"></textarea></div>
                    <div class="metadata-field full-width"><div class="field-label">Keywords <span class="field-tag tag-iptc">IPTC</span></div><textarea class="field-input" id="iptc_keywords"></textarea></div>
                    <div class="metadata-field"><div class="field-label">Date Created <span class="field-tag tag-iptc">IPTC</span></div><input type="date" class="field-input" id="iptc_date_created"></div>
                    <div class="metadata-field"><div class="field-label">By-line (Creator) <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_byline" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Credit <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_credit" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Copyright Notice <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_copyright_notice"></div>
                    <div class="metadata-field"><div class="field-label">Contact <span class="field-tag tag-iptc">IPTC</span></div><input type="text" class="field-input" id="iptc_contact" value="groundswell.co"></div>
                </div>
            </div>

            <div class="metadata-panel" id="panel-xmp">
                <div class="panel-header"><h3>XMP - Extensible Metadata Platform</h3><p>Adobe's standard for metadata in digital files</p></div>
                <div class="metadata-grid">
                    <div class="metadata-field"><div class="field-label">Title <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_title"></div>
                    <div class="metadata-field"><div class="field-label">Label <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_label"></div>
                    <div class="metadata-field"><div class="field-label">Headline <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_headline"></div>
                    <div class="metadata-field full-width"><div class="field-label">Description <span class="field-tag tag-xmp">XMP</span></div><textarea class="field-input" id="xmp_description"></textarea></div>
                    <div class="metadata-field full-width"><div class="field-label">Subject (Keywords) <span class="field-tag tag-xmp">XMP</span></div><textarea class="field-input" id="xmp_subject"></textarea></div>
                    <div class="metadata-field"><div class="field-label">Date Created <span class="field-tag tag-xmp">XMP</span></div><input type="date" class="field-input" id="xmp_date_created"></div>
                    <div class="metadata-field"><div class="field-label">Creator <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_creator" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Credit <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_credit" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Rights/Copyright <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_rights"></div>
                    <div class="metadata-field"><div class="field-label">Copyright Status <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_copyright_status" value="Protected"></div>
                    <div class="metadata-field"><div class="field-label">Source <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_source" value="GroundSwell"></div>
                    <div class="metadata-field"><div class="field-label">Category <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_category" value="Business Ownership Platform"></div>
                    <div class="metadata-field"><div class="field-label">Website <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_website" value="www.groundswell.co"></div>
                    <div class="metadata-field"><div class="field-label">Phone <span class="field-tag tag-xmp">XMP</span></div><input type="text" class="field-input" id="xmp_phone" value="435-214-2997"></div>
                </div>
            </div>

            <button class="new-upload-btn" id="newUploadBtn">Upload Another Image</button>
        </div>

        <section class="marketing">
            <div class="marketing-section">
                <h2>How It Works</h2>
                <p class="section-lead">From raw image to publish-ready asset in three steps.</p>
                <div class="steps-grid">
                    <div class="step-card"><div class="step-number">STEP 01</div><h3>Upload</h3><p>Drop your image — or a specific variant (social, featured, thumbnail). Add a URL or document for extra context.</p></div>
                    <div class="step-card"><div class="step-number">STEP 02</div><h3>Analyze</h3><p>Claude Sonnet 4 reads the visual content and generates 28 metadata fields aligned with GroundSwell's brand language.</p></div>
                    <div class="step-card"><div class="step-number">STEP 03</div><h3>Download</h3><p>Edit any field, then export with metadata baked into the file. Single image or a zipped bundle of every variant.</p></div>
                </div>
            </div>
            <div class="marketing-section">
                <h2>Why Teams Use It</h2>
                <p class="section-lead">One tool, three metadata standards, and a brand voice that stays consistent across every asset.</p>
                <div class="features-grid">
                    <div class="feature-card"><div class="feature-icon">📝</div><h4>Three Standards, One Pass</h4><p>Writes EXIF, IPTC, and XMP simultaneously — the formats used by every major photo library, CMS, and search engine.</p></div>
                    <div class="feature-card"><div class="feature-icon">🎯</div><h4>Brand-Aware AI</h4><p>Trained on GroundSwell concepts: Organizational Alpha, Fractal Flywheel℠, Changing Pains℠, Direct Ownership.</p></div>
                    <div class="feature-card"><div class="feature-icon">📐</div><h4>Built-In Variants</h4><p>Content (original) · Social (1200×720) · Featured (700×400) · Thumbnail (232×245) — tagged and resized automatically.</p></div>
                    <div class="feature-card"><div class="feature-icon">🔗</div><h4>Context Reference</h4><p>Point the AI at a URL, PDF, Word, Excel, or CSV — it reads the source to produce sharper, more accurate metadata.</p></div>
                    <div class="feature-card"><div class="feature-icon">✏️</div><h4>Every Field Editable</h4><p>AI suggestions are a starting point, not a verdict. Tweak titles, keywords, and descriptions before export.</p></div>
                    <div class="feature-card"><div class="feature-icon">📦</div><h4>Bulk Export</h4><p>Download a single variant or grab every variant in one zip — filenames follow the brand's naming convention automatically.</p></div>
                </div>
            </div>
            <div class="marketing-section">
                <h2>Brand Defaults, Baked In</h2>
                <p class="section-lead">Every export ships with the GroundSwell identity already written into the file.</p>
                <div class="brand-defaults">
                    <div class="defaults-grid">
                        <div class="default-row"><span class="default-key">Artist / Creator</span><span class="default-value">GroundSwell</span></div>
                        <div class="default-row"><span class="default-key">Credit</span><span class="default-value">GroundSwell</span></div>
                        <div class="default-row"><span class="default-key">Copyright</span><span class="default-value">Copyright <span id="copyrightYear"></span> GroundSwell</span></div>
                        <div class="default-row"><span class="default-key">Copyright Status</span><span class="default-value">Protected</span></div>
                        <div class="default-row"><span class="default-key">Category</span><span class="default-value">Business Ownership Platform</span></div>
                        <div class="default-row"><span class="default-key">Contact Website</span><span class="default-value">www.groundswell.co</span></div>
                        <div class="default-row"><span class="default-key">Contact Phone</span><span class="default-value">435-214-2997</span></div>
                        <div class="default-row"><span class="default-key">Contact (IPTC)</span><span class="default-value">groundswell.co</span></div>
                    </div>
                </div>
            </div>
            <div class="marketing-section" id="pricing">
                <h2>Plans & Pricing</h2>
                <p class="section-lead">Pick a monthly subscription, or pay per image with credit bundles. Demo pricing — payments launching soon.</p>
                <div class="pricing-grid">
                    <div class="plan-card">
                        <div class="plan-name">Starter</div>
                        <div class="plan-price">$0<span class="per"> / mo</span></div>
                        <p class="plan-desc">Try the tool on a few images per month.</p>
                        <ul class="plan-features">
                            <li>5 images / month</li>
                            <li>EXIF + IPTC metadata</li>
                            <li>Single-variant export</li>
                        </ul>
                        <button class="btn btn-secondary plan-cta" data-plan="starter">Current plan</button>
                    </div>
                    <div class="plan-card featured">
                        <div class="plan-badge">MOST POPULAR</div>
                        <div class="plan-name">Pro</div>
                        <div class="plan-price">$29<span class="per"> / mo</span></div>
                        <p class="plan-desc">For marketing teams shipping content weekly.</p>
                        <ul class="plan-features">
                            <li>200 images / month</li>
                            <li>EXIF + IPTC + XMP</li>
                            <li>All 4 variants (content, social, featured, thumb)</li>
                            <li>Context reference uploads</li>
                            <li>Priority Claude Sonnet 4 queue</li>
                        </ul>
                        <button class="btn btn-primary plan-cta" data-plan="pro">Add to cart</button>
                    </div>
                    <div class="plan-card">
                        <div class="plan-name">Team</div>
                        <div class="plan-price">$99<span class="per"> / mo</span></div>
                        <p class="plan-desc">Multi-seat access with API automation.</p>
                        <ul class="plan-features">
                            <li>1,000 images / month</li>
                            <li>Everything in Pro</li>
                            <li>5 team seats</li>
                            <li>REST API access</li>
                            <li>Dedicated onboarding</li>
                        </ul>
                        <button class="btn btn-secondary plan-cta" data-plan="team">Add to cart</button>
                    </div>
                </div>
                <div class="credits-card">
                    <div class="credits-info">
                        <h3>Pay per image</h3>
                        <p>No subscription? Buy image credits in bundles. One credit = one image with full metadata. Credits never expire.</p>
                    </div>
                    <div class="credits-packs">
                        <button class="pack-btn" data-credits="10"><span class="pack-qty">10 credits</span><span class="pack-price">$6 · $0.60/img</span></button>
                        <button class="pack-btn" data-credits="50"><span class="pack-qty">50 credits</span><span class="pack-price">$25 · $0.50/img</span></button>
                        <button class="pack-btn" data-credits="200"><span class="pack-qty">200 credits</span><span class="pack-price">$80 · $0.40/img</span></button>
                    </div>
                </div>
            </div>
            <div class="cta-band">
                <h2>Ready to tag your next image?</h2>
                <p>Scroll up and drop a file. You'll have publish-ready metadata in under a minute.</p>
                <button class="btn btn-primary" type="button" onclick="window.scrollTo({top:0,behavior:'smooth'})">Start Uploading</button>
            </div>
        </section>

        <footer class="brand-footer">
            <div class="footer-grid">
                <div class="footer-brand">
                    <div class="logo-small">GroundSwell℠</div>
                    <p>A Business Ownership Platform creating Organizational Alpha through incremental returns and decreased risks.</p>
                </div>
                <div class="footer-col">
                    <h5>Product</h5>
                    <a href="#uploadSection">Metadata Tool</a>
                    <a href="https://groundswell.co" target="_blank" rel="noopener">GroundSwell.co</a>
                </div>
                <div class="footer-col">
                    <h5>Contact</h5>
                    <a href="https://groundswell.co" target="_blank" rel="noopener">www.groundswell.co</a>
                    <a href="tel:4352142997">435-214-2997</a>
                </div>
            </div>
            <div class="footer-bottom">Powered by Claude AI · © <span id="year"></span> <a href="https://groundswell.co" target="_blank" rel="noopener">GroundSwell℠</a> — All rights reserved</div>
        </footer>
    </div>

    <button class="cart-fab" id="cartFab" aria-label="Open cart">🛒<span class="cart-badge" id="cartBadge">0</span></button>
    <div class="cart-backdrop" id="cartBackdrop"></div>
    <aside class="cart-drawer" id="cartDrawer" aria-label="Shopping cart">
        <div class="cart-header">
            <h3>Your Cart</h3>
            <button class="cart-close" id="cartClose" aria-label="Close cart">×</button>
        </div>
        <div class="cart-body" id="cartBody">
            <div class="cart-empty" id="cartEmpty">
                <div class="cart-empty-icon">🛒</div>
                <p>Your cart is empty</p>
                <p style="font-size:0.85rem;margin-top:6px">Add a subscription or image credits to get started.</p>
            </div>
        </div>
        <div class="cart-footer">
            <div class="cart-totals">
                <div class="cart-row"><span>Subtotal</span><span id="cartSubtotal">$0.00</span></div>
                <div class="cart-row"><span>Tax</span><span>Calculated at checkout</span></div>
                <div class="cart-row total"><span>Total</span><span id="cartTotal">$0.00</span></div>
            </div>
            <button class="btn btn-primary checkout-btn" id="checkoutBtn" disabled>Checkout</button>
        </div>
    </aside>
    <div class="modal-backdrop" id="checkoutModal">
        <div class="modal">
            <div class="modal-icon">🚀</div>
            <h3>Demo checkout</h3>
            <p>In production, this button opens Stripe Checkout. For the demo we stop here — your cart is ready to hand off to the payments integration.</p>
            <button class="btn btn-primary" id="modalClose">Got it</button>
        </div>
    </div>

    <div class="saving-overlay" id="savingOverlay"><div><div class="spinner"></div><p>Saving metadata...</p></div></div>

    <script>
        document.getElementById('year').textContent = new Date().getFullYear();
        const copyrightYearEl = document.getElementById('copyrightYear');
        if (copyrightYearEl) copyrightYearEl.textContent = new Date().getFullYear();
        const uploadSection = document.getElementById('uploadSection');
        const fileInput = document.getElementById('fileInput');
        const processing = document.getElementById('processing');
        const results = document.getElementById('results');
        const statusMessage = document.getElementById('statusMessage');
        const savingOverlay = document.getElementById('savingOverlay');

        let currentFileId = null;
        let originalMetadata = null;
        let selectedFile = null;
        let selectedSocialFile = null;
        let selectedFeaturedFile = null;
        let selectedThumbnailFile = null;
        let uploadMode = 'full';
        let singleVariantType = null;
        const variantLabels = { content: 'Content (full size)', social: 'Social (1200 x 720)', featured: 'Featured (700 x 400)', thumbnail: 'Thumbnail (232 x 245)' };

        // Context Reference toggle
        document.getElementById('contextHeader').addEventListener('click', () => {
            const body = document.getElementById('contextBody');
            const toggle = document.getElementById('contextToggle');
            body.classList.toggle('open');
            toggle.innerHTML = body.classList.contains('open') ? '&#9650;' : '&#9660;';
        });

        // PDF file picker
        document.getElementById('pdfBtn').addEventListener('click', (e) => { e.stopPropagation(); document.getElementById('pdfInput').click(); });
        document.getElementById('pdfInput').addEventListener('change', (e) => {
            const name = e.target.files[0] ? e.target.files[0].name : 'No file selected';
            document.getElementById('pdfFileName').textContent = name;
            updateContextBadge();
        });
        document.getElementById('contextUrl').addEventListener('input', updateContextBadge);

        function updateContextBadge() {
            const hasUrl = document.getElementById('contextUrl').value.trim() !== '';
            const hasPdf = document.getElementById('pdfInput').files.length > 0;
            document.getElementById('contextBadge').style.display = (hasUrl || hasPdf) ? 'inline-block' : 'none';
        }

        // Variant select cards (single variant upload)
        document.querySelectorAll('.variant-select-card').forEach(card => {
            const variant = card.dataset.variant;
            const input = card.querySelector('input[type="file"]');
            card.addEventListener('click', () => input.click());
            input.addEventListener('change', (e) => {
                if (e.target.files.length) { uploadMode = 'single'; singleVariantType = variant; handleFile(e.target.files[0]); }
            });
        });

        // Variant image uploads (social & featured)
        function setupVariantUpload(type) {
            const zone = document.getElementById(type + 'UploadZone');
            const input = document.getElementById(type + 'Input');
            const preview = document.getElementById(type + 'Preview');
            const placeholder = document.getElementById(type + 'Placeholder');
            const removeBtn = document.getElementById(type + 'RemoveBtn');

            zone.addEventListener('click', () => input.click());
            zone.addEventListener('dragover', (e) => { e.preventDefault(); zone.style.borderColor = '#7FBBE6'; });
            zone.addEventListener('dragleave', () => { zone.style.borderColor = ''; });
            zone.addEventListener('drop', (e) => { e.preventDefault(); zone.style.borderColor = ''; if (e.dataTransfer.files.length) loadVariant(e.dataTransfer.files[0], type); });
            input.addEventListener('change', (e) => { if (e.target.files.length) loadVariant(e.target.files[0], type); });

            removeBtn.addEventListener('click', (e) => {
                e.stopPropagation();
                if (type === 'social') selectedSocialFile = null; else if (type === 'featured') selectedFeaturedFile = null; else selectedThumbnailFile = null;
                preview.style.display = 'none'; placeholder.style.display = 'block'; removeBtn.style.display = 'none';
                zone.classList.remove('has-file'); input.value = '';
            });
        }
        function loadVariant(file, type) {
            if (!file.type.startsWith('image/')) return;
            if (type === 'social') selectedSocialFile = file; else if (type === 'featured') selectedFeaturedFile = file; else selectedThumbnailFile = file;
            const reader = new FileReader();
            const preview = document.getElementById(type + 'Preview');
            const placeholder = document.getElementById(type + 'Placeholder');
            const removeBtn = document.getElementById(type + 'RemoveBtn');
            const zone = document.getElementById(type + 'UploadZone');
            reader.onload = (e) => { preview.src = e.target.result; preview.style.display = 'block'; placeholder.style.display = 'none'; removeBtn.style.display = 'inline-block'; zone.classList.add('has-file'); };
            reader.readAsDataURL(file);
        }
        setupVariantUpload('social');
        setupVariantUpload('featured');
        setupVariantUpload('thumbnail');

        // Change Image button
        document.getElementById('changeImageBtn').addEventListener('click', () => {
            selectedFile = null; uploadMode = 'full'; singleVariantType = null;
            document.getElementById('previewSection').style.display = 'none';
            uploadSection.style.display = 'block';
            document.getElementById('variantSelectSection').style.display = 'block';
            fileInput.value = '';
        });

        // Generate Metadata button
        document.getElementById('generateBtn').addEventListener('click', async () => {
            if (!selectedFile) return;
            document.getElementById('previewSection').style.display = 'none';
            processing.style.display = 'block';
            results.style.display = 'none';

            const formData = new FormData();
            formData.append('file', selectedFile);
            if (uploadMode === 'single' && singleVariantType) {
                formData.append('variant_type', singleVariantType);
            } else {
                if (selectedSocialFile) formData.append('social_file', selectedSocialFile);
                if (selectedFeaturedFile) formData.append('featured_file', selectedFeaturedFile);
                if (selectedThumbnailFile) formData.append('thumbnail_file', selectedThumbnailFile);
            }
            const contextUrl = document.getElementById('contextUrl').value.trim();
            if (contextUrl) formData.append('context_url', contextUrl);
            const pdfInput = document.getElementById('pdfInput');
            if (pdfInput.files.length > 0) formData.append('context_file', pdfInput.files[0]);

            try {
                const response = await fetch('/upload', { method: 'POST', body: formData });
                const data = await response.json();
                if (!response.ok) throw new Error(data.detail || 'Upload failed');
                currentFileId = data.file_id;
                originalMetadata = data.metadata;
                document.getElementById('resultTitle').textContent = data.metadata.title || 'Image Metadata';
                populateFields(data.metadata);
                results.style.display = 'block';
            } catch (error) {
                alert('Error: ' + error.message);
                document.getElementById('previewSection').style.display = 'block';
            } finally {
                processing.style.display = 'none';
            }
        });

        document.querySelectorAll('.tab-btn').forEach(btn => {
            btn.addEventListener('click', () => {
                document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
                document.querySelectorAll('.metadata-panel').forEach(p => p.classList.remove('active'));
                btn.classList.add('active');
                document.getElementById('panel-' + btn.dataset.tab).classList.add('active');
            });
        });

        uploadSection.addEventListener('click', () => { uploadMode = 'full'; singleVariantType = null; fileInput.click(); });
        uploadSection.addEventListener('dragover', e => { e.preventDefault(); uploadSection.style.borderColor = '#7FBBE6'; });
        uploadSection.addEventListener('dragleave', () => { uploadSection.style.borderColor = 'rgba(255,255,255,0.2)'; });
        uploadSection.addEventListener('drop', e => { e.preventDefault(); uploadSection.style.borderColor = 'rgba(255,255,255,0.2)'; uploadMode = 'full'; singleVariantType = null; if (e.dataTransfer.files.length) handleFile(e.dataTransfer.files[0]); });
        fileInput.addEventListener('change', e => { if (e.target.files.length) { if (uploadMode !== 'single') { uploadMode = 'full'; singleVariantType = null; } handleFile(e.target.files[0]); } });

        document.getElementById('newUploadBtn').addEventListener('click', () => {
            results.style.display = 'none'; document.getElementById('previewSection').style.display = 'none'; uploadSection.style.display = 'block'; document.getElementById('variantSelectSection').style.display = 'block'; fileInput.value = ''; selectedFile = null; selectedSocialFile = null; selectedFeaturedFile = null; selectedThumbnailFile = null; currentFileId = null; originalMetadata = null; statusMessage.className = 'status-message'; uploadMode = 'full'; singleVariantType = null;
            document.getElementById('contextUrl').value = ''; document.getElementById('pdfInput').value = ''; document.getElementById('pdfFileName').textContent = 'No file selected'; document.getElementById('contextBadge').style.display = 'none';
            document.getElementById('contextBody').classList.remove('open'); document.getElementById('contextToggle').innerHTML = '&#9660;';
            ['social', 'featured', 'thumbnail'].forEach(t => { document.getElementById(t+'Preview').style.display = 'none'; document.getElementById(t+'Placeholder').style.display = 'block'; document.getElementById(t+'RemoveBtn').style.display = 'none'; document.getElementById(t+'UploadZone').classList.remove('has-file'); document.getElementById(t+'Input').value = ''; });
            document.getElementById('downloadSection').style.display = 'none'; currentDownloadFiles = [];
        });

        document.getElementById('resetBtn').addEventListener('click', () => { if (originalMetadata) { populateFields(originalMetadata); showStatus('Fields reset to AI-generated values', 'success'); } });

        let currentDownloadFiles = [];
        let currentZipName = '';

        document.getElementById('saveDownloadBtn').addEventListener('click', async () => {
            if (!currentFileId) return;
            savingOverlay.classList.add('active');
            try {
                const response = await fetch('/save-metadata', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ file_id: currentFileId, metadata: gatherMetadata() })
                });
                const data = await response.json();
                if (!response.ok) throw new Error(data.detail || 'Save failed');

                currentDownloadFiles = data.files;
                currentZipName = data.zip_name;
                const list = document.getElementById('downloadList');
                list.innerHTML = '';
                data.files.forEach((file, i) => {
                    const item = document.createElement('div');
                    item.className = 'download-item';
                    const typeLabel = file.variant ? file.variant.charAt(0).toUpperCase() + file.variant.slice(1) : '';
                    item.innerHTML = '<span class="download-item-name">' + file.filename + '</span>' +
                        '<span class="download-item-type">' + typeLabel + '</span>' +
                        '<button class="btn btn-secondary btn-small" onclick="downloadSingleFile(' + i + ')">Download</button>';
                    list.appendChild(item);
                });
                document.getElementById('downloadSection').style.display = 'block';
                document.getElementById('downloadAllRow').style.display = data.files.length > 1 ? 'block' : 'none';
                showStatus('Metadata saved! ' + data.files.length + ' file(s) ready for download.', 'success');
            } catch (error) {
                showStatus('Error: ' + error.message, 'error');
            } finally {
                savingOverlay.classList.remove('active');
            }
        });

        function downloadSingleFile(index) {
            const file = currentDownloadFiles[index];
            const link = document.createElement('a');
            link.href = 'data:application/octet-stream;base64,' + file.image_data;
            link.download = file.filename;
            link.click();
        }

        document.getElementById('downloadAllBtn').addEventListener('click', async () => {
            const zip = new JSZip();
            currentDownloadFiles.forEach(file => { zip.file(file.filename, file.image_data, {base64: true}); });
            const blob = await zip.generateAsync({type: 'blob'});
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = currentZipName;
            link.click();
            URL.revokeObjectURL(link.href);
        });

        function handleFile(file) {
            if (!file.type.startsWith('image/')) { alert('Please upload an image file'); return; }
            selectedFile = file;
            const reader = new FileReader();
            reader.onload = (e) => { document.getElementById('imagePreview').src = e.target.result; };
            reader.readAsDataURL(file);
            document.getElementById('previewFilename').textContent = file.name;
            uploadSection.style.display = 'none';
            document.getElementById('variantSelectSection').style.display = 'none';

            if (uploadMode === 'single') {
                document.getElementById('previewVariantLabel').textContent = variantLabels[singleVariantType] || singleVariantType;
                document.querySelector('.variant-uploads-section').style.display = 'none';
            } else {
                document.getElementById('previewVariantLabel').textContent = variantLabels['content'];
                document.querySelector('.variant-uploads-section').style.display = '';
            }
            document.getElementById('previewSection').style.display = 'block';
        }

        function populateFields(meta) {
            const year = (meta.create_date || new Date().toISOString().split('T')[0]).substring(0, 4);
            const dateValue = meta.create_date || new Date().toISOString().split('T')[0];
            const keywords = Array.isArray(meta.keywords) ? meta.keywords.join(', ') : (meta.keywords || '');

            document.getElementById('exif_create_date').value = dateValue;
            document.getElementById('exif_artist').value = meta.artist || 'GroundSwell';
            document.getElementById('exif_copyright').value = meta.copyright || 'Copyright ' + year + ' GroundSwell';
            document.getElementById('exif_description').value = meta.description || '';
            document.getElementById('exif_user_comment').value = keywords;

            document.getElementById('iptc_object_name').value = meta.title || '';
            document.getElementById('iptc_headline').value = meta.headline || '';
            document.getElementById('iptc_caption').value = meta.description || '';
            document.getElementById('iptc_keywords').value = keywords;
            document.getElementById('iptc_date_created').value = dateValue;
            document.getElementById('iptc_byline').value = meta.artist || 'GroundSwell';
            document.getElementById('iptc_credit').value = meta.credit || 'GroundSwell';
            document.getElementById('iptc_copyright_notice').value = meta.copyright || 'Copyright ' + year + ' GroundSwell';
            document.getElementById('iptc_contact').value = meta.contact || 'groundswell.co';

            document.getElementById('xmp_title').value = meta.title || '';
            document.getElementById('xmp_label').value = meta.title || '';
            document.getElementById('xmp_headline').value = meta.headline || '';
            document.getElementById('xmp_description').value = meta.description || '';
            document.getElementById('xmp_subject').value = keywords;
            document.getElementById('xmp_date_created').value = dateValue;
            document.getElementById('xmp_creator').value = meta.artist || 'GroundSwell';
            document.getElementById('xmp_credit').value = meta.credit || 'GroundSwell';
            document.getElementById('xmp_rights').value = meta.copyright || 'Copyright ' + year + ' GroundSwell';
            document.getElementById('xmp_copyright_status').value = 'Protected';
            document.getElementById('xmp_source').value = 'GroundSwell';
            document.getElementById('xmp_category').value = meta.category || 'Business Ownership Platform';
            document.getElementById('xmp_website').value = meta.website || 'www.groundswell.co';
            document.getElementById('xmp_phone').value = meta.phone || '435-214-2997';
        }

        function gatherMetadata() {
            return {
                exif_create_date: document.getElementById('exif_create_date').value,
                exif_artist: document.getElementById('exif_artist').value,
                exif_copyright: document.getElementById('exif_copyright').value,
                exif_description: document.getElementById('exif_description').value,
                exif_user_comment: document.getElementById('exif_user_comment').value,
                iptc_object_name: document.getElementById('iptc_object_name').value,
                iptc_headline: document.getElementById('iptc_headline').value,
                iptc_caption: document.getElementById('iptc_caption').value,
                iptc_keywords: document.getElementById('iptc_keywords').value,
                iptc_date_created: document.getElementById('iptc_date_created').value,
                iptc_byline: document.getElementById('iptc_byline').value,
                iptc_credit: document.getElementById('iptc_credit').value,
                iptc_copyright_notice: document.getElementById('iptc_copyright_notice').value,
                iptc_contact: document.getElementById('iptc_contact').value,
                xmp_title: document.getElementById('xmp_title').value,
                xmp_label: document.getElementById('xmp_label').value,
                xmp_headline: document.getElementById('xmp_headline').value,
                xmp_description: document.getElementById('xmp_description').value,
                xmp_subject: document.getElementById('xmp_subject').value,
                xmp_date_created: document.getElementById('xmp_date_created').value,
                xmp_creator: document.getElementById('xmp_creator').value,
                xmp_credit: document.getElementById('xmp_credit').value,
                xmp_rights: document.getElementById('xmp_rights').value,
                xmp_copyright_status: document.getElementById('xmp_copyright_status').value,
                xmp_source: document.getElementById('xmp_source').value,
                xmp_category: document.getElementById('xmp_category').value,
                xmp_website: document.getElementById('xmp_website').value,
                xmp_phone: document.getElementById('xmp_phone').value
            };
        }

        function showStatus(message, type) { statusMessage.textContent = message; statusMessage.className = 'status-message ' + type; }
    </script>

    <script>
    (function() {
        const cart = { plan: null, credits: 0 };
        const PLANS = {
            pro: { name: 'Pro subscription', price: 29, meta: 'Monthly · 200 images' },
            team: { name: 'Team subscription', price: 99, meta: 'Monthly · 1,000 images · 5 seats' }
        };
        function creditPrice(qty) {
            if (qty >= 200) return qty * 0.40;
            if (qty >= 50) return qty * 0.50;
            return qty * 0.60;
        }
        const fab = document.getElementById('cartFab');
        const badge = document.getElementById('cartBadge');
        const drawer = document.getElementById('cartDrawer');
        const backdrop = document.getElementById('cartBackdrop');
        const body = document.getElementById('cartBody');
        const empty = document.getElementById('cartEmpty');
        const subtotalEl = document.getElementById('cartSubtotal');
        const totalEl = document.getElementById('cartTotal');
        const checkoutBtn = document.getElementById('checkoutBtn');
        const modal = document.getElementById('checkoutModal');

        function openDrawer() { drawer.classList.add('open'); backdrop.classList.add('open'); }
        function closeDrawer() { drawer.classList.remove('open'); backdrop.classList.remove('open'); }
        fab.addEventListener('click', openDrawer);
        document.getElementById('cartClose').addEventListener('click', closeDrawer);
        backdrop.addEventListener('click', closeDrawer);

        function render() {
            body.querySelectorAll('.cart-item').forEach(n => n.remove());
            const hasItems = cart.plan || cart.credits > 0;
            empty.style.display = hasItems ? 'none' : 'block';

            if (cart.plan) {
                const p = PLANS[cart.plan];
                const node = document.createElement('div');
                node.className = 'cart-item';
                node.innerHTML =
                    '<div class="cart-item-top">' +
                    '<div><div class="cart-item-name">' + p.name + '</div><div class="cart-item-meta">' + p.meta + '</div></div>' +
                    '<button class="cart-item-remove" data-remove="plan" aria-label="Remove">×</button>' +
                    '</div>' +
                    '<div class="cart-item-price">$' + p.price.toFixed(2) + ' / mo</div>';
                body.appendChild(node);
            }
            if (cart.credits > 0) {
                const price = creditPrice(cart.credits);
                const rate = (price / cart.credits).toFixed(2);
                const node = document.createElement('div');
                node.className = 'cart-item';
                node.innerHTML =
                    '<div class="cart-item-top">' +
                    '<div><div class="cart-item-name">Image credits</div><div class="cart-item-meta">$' + rate + ' / image · volume pricing</div></div>' +
                    '<button class="cart-item-remove" data-remove="credits" aria-label="Remove">×</button>' +
                    '</div>' +
                    '<div style="display:flex;justify-content:space-between;align-items:center">' +
                    '<div class="cart-qty"><button data-credit-delta="-10" aria-label="Decrease">−</button><span>' + cart.credits + '</span><button data-credit-delta="10" aria-label="Increase">+</button></div>' +
                    '<div class="cart-item-price">$' + price.toFixed(2) + '</div>' +
                    '</div>';
                body.appendChild(node);
            }

            const subtotal = (cart.plan ? PLANS[cart.plan].price : 0) + (cart.credits > 0 ? creditPrice(cart.credits) : 0);
            subtotalEl.textContent = '$' + subtotal.toFixed(2);
            totalEl.textContent = '$' + subtotal.toFixed(2);

            const count = (cart.plan ? 1 : 0) + (cart.credits > 0 ? 1 : 0);
            badge.textContent = count;
            badge.classList.toggle('visible', count > 0);
            checkoutBtn.disabled = count === 0;
        }

        document.querySelectorAll('.plan-cta').forEach(btn => {
            btn.addEventListener('click', () => {
                const plan = btn.dataset.plan;
                if (plan === 'starter') return;
                cart.plan = plan;
                render();
                openDrawer();
            });
        });

        document.querySelectorAll('.pack-btn').forEach(btn => {
            btn.addEventListener('click', () => {
                cart.credits += parseInt(btn.dataset.credits, 10);
                render();
                openDrawer();
            });
        });

        body.addEventListener('click', (e) => {
            const removeTarget = e.target.closest('[data-remove]');
            if (removeTarget) {
                const key = removeTarget.dataset.remove;
                if (key === 'plan') cart.plan = null;
                if (key === 'credits') cart.credits = 0;
                render();
                return;
            }
            const deltaBtn = e.target.closest('[data-credit-delta]');
            if (deltaBtn) {
                const d = parseInt(deltaBtn.dataset.creditDelta, 10);
                cart.credits = Math.max(0, cart.credits + d);
                render();
            }
        });

        checkoutBtn.addEventListener('click', () => { modal.classList.add('open'); });
        document.getElementById('modalClose').addEventListener('click', () => {
            modal.classList.remove('open');
            closeDrawer();
        });
        modal.addEventListener('click', (e) => { if (e.target === modal) modal.classList.remove('open'); });

        render();
    })();
    </script>
</body>
</html>'''


@app.get("/", response_class=HTMLResponse)
async def home():
    return HTML_TEMPLATE


FONTS_DIR = Path(__file__).parent / "fonts"
ALLOWED_FONTS = {"NHaasGroteskTXPro-65Md.ttf", "NeuzeitGro-Reg.ttf"}


@app.get("/fonts/{filename}")
async def serve_font(filename: str):
    if filename not in ALLOWED_FONTS:
        raise HTTPException(status_code=404, detail="Font not found")
    font_path = FONTS_DIR / filename
    if not font_path.is_file():
        raise HTTPException(status_code=404, detail="Font not found")
    return Response(
        content=font_path.read_bytes(),
        media_type="font/ttf",
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )


@app.post("/upload")
async def upload_image(
    file: UploadFile = File(...),
    social_file: Optional[UploadFile] = File(None),
    featured_file: Optional[UploadFile] = File(None),
    thumbnail_file: Optional[UploadFile] = File(None),
    variant_type: Optional[str] = Form(None),
    context_url: Optional[str] = Form(None),
    context_file: Optional[UploadFile] = File(None),
):
    """Upload and analyze image with AI."""

    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type")

    file_id = str(uuid.uuid4())[:8]
    original_filename = file.filename or "image.jpg"

    content = await file.read()

    # Read optional variant files
    social_data = None
    social_filename = None
    if social_file and social_file.filename and social_file.size and social_file.size > 0:
        social_data = await social_file.read()
        social_filename = social_file.filename

    featured_data = None
    featured_filename = None
    if featured_file and featured_file.filename and featured_file.size and featured_file.size > 0:
        featured_data = await featured_file.read()
        featured_filename = featured_file.filename

    thumbnail_data = None
    thumbnail_filename = None
    if thumbnail_file and thumbnail_file.filename and thumbnail_file.size and thumbnail_file.size > 0:
        thumbnail_data = await thumbnail_file.read()
        thumbnail_filename = thumbnail_file.filename

    # Get reference context if provided
    reference_context = ""
    if context_url and context_url.strip():
        reference_context = fetch_url_content(context_url.strip())
    elif context_file and context_file.filename and context_file.size and context_file.size > 0:
        file_data = await context_file.read()
        reference_context = extract_file_text(file_data, context_file.filename)

    # Store all files in memory
    file_storage[file_id] = {
        "data": content,
        "filename": original_filename,
        "social_data": social_data,
        "social_filename": social_filename,
        "featured_data": featured_data,
        "featured_filename": featured_filename,
        "thumbnail_data": thumbnail_data,
        "thumbnail_filename": thumbnail_filename,
        "variant_type": variant_type or "content",
    }

    try:
        metadata = await analyze_image_with_claude(content, original_filename, reference_context)
        year = metadata.get("create_date", datetime.now().strftime("%Y-%m-%d"))[:4]

        return JSONResponse({
            "success": True,
            "file_id": file_id,
            "metadata": {
                "title": metadata.get("title", ""),
                "headline": metadata.get("headline", ""),
                "description": metadata.get("description", ""),
                "keywords": metadata.get("keywords", []),
                "category": metadata.get("category", "Business Ownership Platform"),
                "supplemental_category": metadata.get("supplemental_category", ""),
                "create_date": metadata.get("create_date", datetime.now().strftime("%Y-%m-%d")),
                "artist": "GroundSwell",
                "copyright": f"Copyright {year} GroundSwell",
                "credit": "GroundSwell",
                "contact": "groundswell.co",
                "website": "www.groundswell.co",
                "phone": "435-214-2997"
            }
        })

    except Exception as e:
        if file_id in file_storage:
            del file_storage[file_id]
        raise HTTPException(status_code=500, detail=f"Error processing image: {str(e)}")


@app.post("/save-metadata")
async def save_metadata(request: SaveMetadataRequest):
    """Save metadata to all uploaded image variants and return as base64 (single or zip)."""

    file_id = request.file_id
    metadata = request.metadata

    if file_id not in file_storage:
        raise HTTPException(status_code=404, detail="File not found. Please upload again.")

    stored = file_storage[file_id]
    content_data = stored["data"]
    content_filename = stored["filename"]
    date_str = metadata.get('exif_create_date', datetime.now().strftime('%Y-%m-%d'))

    try:
        # Build list of variants to process: (variant_name, image_bytes, file_ext)
        vtype = stored.get("variant_type", "content")
        has_extra_variants = stored.get("social_data") or stored.get("featured_data") or stored.get("thumbnail_data")

        if has_extra_variants or vtype == "content":
            # Full mode: content + any extra variants
            variants_to_process = [("content", content_data, Path(content_filename).suffix)]
            if stored.get("social_data"):
                variants_to_process.append(("social", stored["social_data"], Path(stored["social_filename"]).suffix))
            if stored.get("featured_data"):
                variants_to_process.append(("featured", stored["featured_data"], Path(stored["featured_filename"]).suffix))
            if stored.get("thumbnail_data"):
                variants_to_process.append(("thumbnail", stored["thumbnail_data"], Path(stored["thumbnail_filename"]).suffix))
        else:
            # Single variant mode: just the selected variant type
            variants_to_process = [(vtype, content_data, Path(content_filename).suffix)]

        processed_files = []
        for variant_name, img_data, ext in variants_to_process:
            # Resize to target dimensions (content stays full size)
            dims = VARIANT_DIMENSIONS.get(variant_name)
            if dims:
                img_data = resize_image_to_fit(img_data, dims[0], dims[1], ext)

            # Build a fake filename with correct extension for metadata processing
            temp_filename = f"image{ext}"
            final_data = process_image_metadata(img_data, temp_filename, metadata)
            # Use content name but variant's own extension
            variant_fn = Path(content_filename).stem + ext.lower()
            dl_filename = generate_download_filename(variant_fn, variant_name, date_str)
            processed_files.append((dl_filename, final_data))

        cleaned_stem = re.sub(r'[^a-z0-9]+', '-', Path(content_filename).stem.lower()).strip('-')
        zip_name = f"{date_str.replace('-', '.')}-{cleaned_stem}-images.zip"

        return JSONResponse({
            "success": True,
            "files": [
                {"filename": name, "image_data": base64.b64encode(data).decode('utf-8'), "variant": variant_name}
                for (name, data), (variant_name, _, _) in zip(processed_files, variants_to_process)
            ],
            "zip_name": zip_name
        })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving metadata: {str(e)}")
